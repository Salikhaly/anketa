# -*- coding: utf-8 -*-
"""
app.py — PKB Анкета v18  |  Premium UI/UX
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import logging, os, re, sys, json, shutil
from datetime import datetime, timedelta
from pathlib import Path
from threading import Thread
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from main import (
    extract_credit_report, extract_any_credit_report, parse_pension_pdf_tables,
    write_output_legacy_excel, fill_anketa_from_pension,
    calc_pension_avg_vals, validate_credit_data,
    validate_pension_data, mask_iin,
    group_pension_by_org, PensionRow,
)

LOGGER = logging.getLogger("pdf_extract")
LOGGER.setLevel(logging.DEBUG)

SETTINGS_FILE = Path(__file__).parent / 'settings.json'
DB_FILE       = Path(__file__).parent / 'база_анкет.xlsx'

DEFAULT_SETTINGS = {
    "managers": [{"id": f"manager{i}", "name": f"Менеджер {i}"} for i in range(1, 11)],
    "backup_folder": "", "font_size": 10, "recent_files": [],
    "crm_url": "", "crm_key": "",   # автопередача лидов в CRM (Настройки → CRM)
    # Подписка «по времени» (тот же ключ/срок, что и веб-сервис ОПВ).
    # license_url пустой = проверка ОТКЛЮЧЕНА (программа работает как раньше).
    # Впишите сюда URL веб-приложения Apps Script (…/exec), чтобы включить.
    "license_url": "", "access_key": "", "device_id": "",
    "license_ok_once": False, "license_until": None, "license_name": "",
    "license_token": "", "license_sig": "",
}

# ── Дизайн-система ────────────────────────────────────────────────────────────
T = {
    "bg":           "#F5F0E8",
    "bg2":          "#EDE7D9",
    "surface":      "#FFFFFF",
    "surface_warm": "#FBF8F3",
    "wine":         "#6B2737",
    "wine_dark":    "#4E1C28",
    "wine_light":   "#F5E8EB",
    "gold":         "#B5860D",
    "gold_light":   "#FDF6E3",
    "ink":          "#1C1917",
    "ink2":         "#57534E",
    "ink3":         "#A8A29E",
    "ink4":         "#D6D3D1",
    "ok":           "#166534",
    "ok_bg":        "#F0FDF4",
    "warn":         "#92400E",
    "warn_bg":      "#FFFBEB",
    "err":          "#991B1B",
    "err_bg":       "#FEF2F2",
    "border":       "#E7E0D4",
    "border_focus": "#6B2737",
    "font":         "Segoe UI",
}

def c(key): return T[key]

# ── Утилиты ───────────────────────────────────────────────────────────────────
def load_settings():
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, encoding='utf-8') as f:
                d = json.load(f)
            for k, v in DEFAULT_SETTINGS.items():
                d.setdefault(k, v)
            return d
        except: pass
    return dict(DEFAULT_SETTINGS)

def save_settings(s):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
    except Exception as e:
        LOGGER.warning("Настройки: %s", e)

def calc_pension_avg(rows):
    amounts = [r.amount for r in rows if isinstance(r, PensionRow) and r.amount]
    return {"avg_salary": calc_pension_avg_vals(amounts) if amounts else 0}

def safe_fn(t):
    if not t: return "UNKNOWN"
    return re.sub(r'[^\w\s-]', '_', t).strip().replace(' ', '_')

def mgr_folder(n):
    return re.sub(r'[^\w\s-]', '_', n).strip().replace(' ', '_')

# ── База данных ───────────────────────────────────────────────────────────────
DB_HDR = ['Телефон','Дата','Менеджер','ФИО','ИИН','ПКР',
          'Нагрузка/мес','СРЗП','Цель','Нал. на руках','Депозит Н','БВ','ГП','Файл']

def update_db(rec):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        if DB_FILE.exists():
            wb = openpyxl.load_workbook(str(DB_FILE)); ws = wb.active
        else:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'База'
            for ci, h in enumerate(DB_HDR, 1):
                cell = ws.cell(1, ci, h)
                cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
                cell.fill      = PatternFill('solid', fgColor='6B2737')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22
            for ci, w in enumerate([14,16,18,26,16,7,16,12,22,16,12,12,12,38], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        row = ws.max_row + 1
        for ci, v in enumerate([
            rec.get('phone',''), rec.get('date',''), rec.get('manager',''),
            rec.get('fio',''), rec.get('iin',''), rec.get('pkr',''),
            rec.get('load',''), rec.get('srzp',''), rec.get('goal',''),
            rec.get('onhand',''), rec.get('cash',''), rec.get('bv',''), rec.get('gp',''),
            rec.get('path','')], 1):
            cell = ws.cell(row, ci, v)
            cell.font      = Font(name='Calibri', size=10)
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FBF8F3')
        wb.save(str(DB_FILE))
        LOGGER.info("База: строка %s добавлена", row)
    except Exception as e:
        LOGGER.warning("База: %s", e)

# ── Автопередача лида в CRM ───────────────────────────────────────────────────
# После создания/импорта анкеты клиент сам появляется в веб-CRM (этап «Новый лид»).
# Настройки → CRM: адрес (https://crm-ipo.vercel.app) и API-ключ (env PKB_API_KEY в Vercel).
# Любая ошибка сети НЕ мешает созданию анкеты — просто пишем в лог.
def post_to_crm(settings, rec, credits=None):
    url = (settings.get('crm_url') or '').strip().rstrip('/')
    key = (settings.get('crm_key') or '').strip()
    if not url or not key:
        return None  # интеграция не настроена — тихо пропускаем
    try:
        import json as _json
        import urllib.request
        payload = {
            'fio':    rec.get('fio', ''),
            'iin':    rec.get('iin', ''),
            'phone':  rec.get('phone', ''),
            'srzp':   rec.get('srzp', ''),
            'load':   rec.get('load', ''),
            'goal':   rec.get('goal', ''),
            'pkr':    rec.get('pkr', ''),
            'onhand': rec.get('onhand', ''),
            'cash':   rec.get('cash', ''),
            'manager': rec.get('manager', ''),
        }
        if credits:
            payload['credits'] = credits
        req = urllib.request.Request(
            url + '/api/integration/pkb',
            data=_json.dumps(payload, ensure_ascii=False).encode('utf-8'),
            headers={'Content-Type': 'application/json', 'x-api-key': key},
            method='POST')
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = _json.loads(resp.read().decode('utf-8'))
            LOGGER.info("CRM: %s (id=%s)",
                        'обновлён' if body.get('action') == 'updated' else 'создан лид',
                        body.get('id'))
            return body
    except Exception as e:
        LOGGER.warning("CRM: не удалось отправить лид — %s", e)
        return None

# Контракты ПКБ → формат credits для CRM
def contracts_to_credits(active):
    out = []
    for c in (active or [])[:30]:
        out.append({
            'status': 'active',
            'type': getattr(c, 'financing_type', '') or '',
            'creditor': getattr(c, 'creditor', '') or '',
            'amount': getattr(c, 'contract_amount', None) or '',
            'payment': getattr(c, 'periodic_payment', None) or '',
            'outstanding': getattr(c, 'outstanding', None) or '',
            'overdueDays': getattr(c, 'current_overdue_days', None) or '',
            'overdueAmount': getattr(c, 'current_overdue_amount', None) or '',
        })
    return out

# ── Импорт анкеты (Excel) → запись данных клиента ─────────────────────────────
# Читает заполненную анкету (template.xlsx) по тем же ячейкам, куда пишет
# _write_extra / fill_anketa_from_pension. Учитывает объединённые ячейки.
def _num_from(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return str(int(v))
    m = re.search(r'-?\d[\d\s.,]*', str(v).replace(' ', ' '))
    if not m:
        return ''
    d = re.sub(r'[^\d]', '', m.group(0))
    return d or ''

def read_anketa_xlsx(path):
    """Возвращает dict в формате update_db, извлекая данные клиента из анкеты."""
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import coordinate_to_tuple
    wb = openpyxl.load_workbook(str(path), data_only=True)

    def gc(ws, ref):
        if ws is None:
            return None
        cell = ws[ref]
        if isinstance(cell, MergedCell):
            r, col = coordinate_to_tuple(ref)
            for mc in ws.merged_cells.ranges:
                if mc.min_row <= r <= mc.max_row and mc.min_col <= col <= mc.max_col:
                    return ws.cell(mc.min_row, mc.min_col).value
            return None
        return cell.value

    ank = wb['Анкета'] if 'Анкета' in wb.sheetnames else None
    kr  = wb['Кредиты'] if 'Кредиты' in wb.sheetnames else None
    if ank is None:
        raise ValueError('В файле нет листа «Анкета» — это не анкета ПКБ.')

    def s(ws, ref):
        v = gc(ws, ref)
        return str(v).strip() if v is not None else ''

    fio   = s(ank, 'B5')
    iin_r = re.sub(r'\D', '', s(ank, 'B6'))
    iin   = iin_r if len(iin_r) == 12 else ''
    phone = s(ank, 'F6')
    goal  = s(ank, 'H10')
    onhand = _num_from(gc(ank, 'H13'))
    cash   = _num_from(gc(ank, 'E18'))
    bv     = _num_from(gc(ank, 'H18'))
    gp     = _num_from(gc(ank, 'J18'))
    srzp   = _num_from(gc(ank, 'B43'))

    pkr = ''
    if kr is not None:
        m = re.search(r'ПКР\s*(\d+)', s(kr, 'A1'), re.IGNORECASE)
        pkr = m.group(1) if m else ''
    load = 0
    if kr is not None:
        for r in range(6, 24):
            n = _num_from(gc(kr, f'E{r}'))
            if n:
                load += int(n)

    return {
        'phone': phone, 'date': datetime.now().strftime('%d.%m.%Y %H:%M'),
        'manager': '', 'fio': fio, 'iin': iin, 'pkr': pkr,
        'load': str(load) if load else '', 'srzp': srzp, 'goal': goal,
        'onhand': onhand, 'cash': cash, 'bv': bv, 'gp': gp,
        'path': str(path),
    }

# ── Логгер ────────────────────────────────────────────────────────────────────
class TextHandler(logging.Handler):
    def __init__(self, w):
        super().__init__(); self.w = w
        self.setFormatter(logging.Formatter('%(asctime)s  %(levelname)-7s  %(message)s'))
    def emit(self, record):
        msg = self.format(record) + '\n'
        tag = 'err' if record.levelno >= logging.ERROR else \
              'wrn' if record.levelno >= logging.WARNING else 'inf'
        def _():
            self.w.config(state='normal')
            self.w.insert('end', msg, tag)
            self.w.see('end')
            self.w.config(state='disabled')
        self.w.after(0, _)

# ── UI-компоненты ─────────────────────────────────────────────────────────────

class Card(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=c('surface'),
                         highlightbackground=c('border'),
                         highlightthickness=1, **kw)

class SectionLabel(tk.Label):
    def __init__(self, parent, text, **kw):
        super().__init__(parent, text=text.upper(),
                         font=(c('font'), 8, 'bold'),
                         bg=c('surface'), fg=c('ink3'), **kw)

class PrimaryButton(tk.Button):
    def __init__(self, parent, text, command, fs=10, padx=24, pady=12, **kw):
        super().__init__(parent, text=text, command=command,
                         font=(c('font'), fs, 'bold'),
                         bg=c('wine'), fg='white',
                         activebackground=c('wine_dark'),
                         activeforeground='white',
                         relief='flat', cursor='hand2',
                         padx=padx, pady=pady, bd=0, **kw)

class SecondaryButton(tk.Button):
    def __init__(self, parent, text, command, fs=10, padx=16, pady=10, **kw):
        super().__init__(parent, text=text, command=command,
                         font=(c('font'), fs),
                         bg=c('surface'), fg=c('ink2'),
                         activebackground=c('bg2'),
                         activeforeground=c('ink'),
                         relief='flat', cursor='hand2',
                         highlightbackground=c('border'),
                         highlightthickness=1,
                         padx=padx, pady=pady, bd=0, **kw)

class InputField(tk.Frame):
    def __init__(self, parent, label, placeholder, width=18, fs=10, **kw):
        super().__init__(parent, bg=c('surface'), **kw)
        self.ph = placeholder
        self.var = tk.StringVar()
        self.fs = fs
        tk.Label(self, text=label,
                 font=(c('font'), fs-1, 'bold'),
                 bg=c('surface'), fg=c('ink2')).pack(anchor='w', pady=(0,3))
        self.entry = tk.Entry(self, textvariable=self.var, width=width,
                               font=(c('font'), fs), relief='flat',
                               bg=c('surface_warm'), fg=c('ink3'),
                               insertbackground=c('wine'),
                               highlightbackground=c('border'),
                               highlightcolor=c('border_focus'),
                               highlightthickness=1)
        self.entry.pack(fill='x', ipady=7)
        self.entry.insert(0, placeholder)
        self.entry.bind('<FocusIn>',  self._in)
        self.entry.bind('<FocusOut>', self._out)

    def _in(self, _):
        if self.entry.get() == self.ph:
            self.entry.delete(0, 'end')
            self.entry.config(fg=c('ink'))
        self.entry.config(highlightbackground=c('border_focus'))

    def _out(self, _):
        if not self.entry.get().strip():
            self.entry.insert(0, self.ph)
            self.entry.config(fg=c('ink3'))
        self.entry.config(highlightbackground=c('border'))

    def get(self):
        v = self.var.get().strip()
        return '' if v == self.ph else v

class UploadZone(tk.Frame):
    def __init__(self, parent, title, subtitle, required=True, fs=10, **kw):
        self.req = required; self.fs = fs
        super().__init__(parent, bg=c('surface'),
                         highlightbackground=c('wine') if required else c('border'),
                         highlightthickness=2 if required else 1, **kw)
        self.path_var = tk.StringVar()
        self._build(title, subtitle)
        try:
            import tkinterdnd2 as dnd
            self.drop_target_register(dnd.DND_FILES)
            self.dnd_bind('<<Drop>>', lambda e: self._on_drop(e.data))
        except: pass

    def _build(self, title, subtitle):
        fs = self.fs
        accent = c('wine') if self.req else c('ink4')
        tk.Frame(self, bg=accent, width=4).pack(side='left', fill='y')
        body = tk.Frame(self, bg=c('surface'))
        body.pack(side='left', fill='both', expand=True, padx=16, pady=14)

        top = tk.Frame(body, bg=c('surface')); top.pack(fill='x')
        title_frame = tk.Frame(top, bg=c('surface')); title_frame.pack(side='left', fill='x', expand=True)

        badge_bg = c('wine_light') if self.req else c('bg')
        badge_fg = c('wine')       if self.req else c('ink3')
        badge_tx = '● Обязательно' if self.req else '○ Необязательно'
        tk.Label(title_frame, text=badge_tx, font=(c('font'), 7, 'bold'),
                 bg=badge_bg, fg=badge_fg).pack(anchor='w', pady=(0,3))
        tk.Label(title_frame, text=title, font=(c('font'), fs+1, 'bold'),
                 bg=c('surface'), fg=c('ink')).pack(anchor='w')
        tk.Label(title_frame, text=subtitle, font=(c('font'), fs-1),
                 bg=c('surface'), fg=c('ink3')).pack(anchor='w', pady=(2,0))

        self.browse_btn = tk.Button(top, text='  Выбрать PDF  ',
            font=(c('font'), fs-1, 'bold'),
            bg=c('wine') if self.req else c('surface'),
            fg='white' if self.req else c('ink2'),
            activebackground=c('wine_dark') if self.req else c('bg2'),
            activeforeground='white' if self.req else c('ink'),
            highlightbackground=c('wine') if self.req else c('border'),
            highlightthickness=1, relief='flat', cursor='hand2',
            padx=14, pady=8, command=self._browse)
        self.browse_btn.pack(side='right', padx=(12,0))

        tk.Frame(body, bg=c('border'), height=1).pack(fill='x', pady=(10,8))

        status_row = tk.Frame(body, bg=c('surface')); status_row.pack(fill='x')
        self.icon_lbl = tk.Label(status_row, text='○', font=(c('font'), fs),
                                  bg=c('surface'), fg=c('ink3'))
        self.icon_lbl.pack(side='left')
        self.status_lbl = tk.Label(status_row,
            text=' Файл не выбран — нажмите кнопку или перетащите PDF сюда',
            font=(c('font'), fs-1, 'italic'),
            bg=c('surface'), fg=c('ink3'), anchor='w')
        self.status_lbl.pack(side='left', fill='x', expand=True)
        self.clear_btn = tk.Button(status_row, text='✕',
            font=(c('font'), fs-2, 'bold'),
            bg=c('err_bg'), fg=c('err'),
            activebackground=c('err'), activeforeground='white',
            relief='flat', cursor='hand2', padx=8, pady=3,
            command=self._clear)

    def _on_drop(self, data):
        p = data.strip().strip('{}').split('} {')[0].strip('"')
        if p.lower().endswith('.pdf') and Path(p).exists(): self.set_path(p)
        else: messagebox.showwarning('Неверный файл', 'Перетащите PDF-файл.')

    def _browse(self):
        p = filedialog.askopenfilename(title='Выберите PDF',
            filetypes=[('PDF документы','*.pdf'),('Все файлы','*.*')])
        if p: self.set_path(p)

    def _clear(self):
        self.path_var.set('')
        self.icon_lbl.config(text='○', fg=c('ink3'))
        self.status_lbl.config(
            text=' Файл не выбран — нажмите кнопку или перетащите PDF сюда',
            fg=c('ink3'), font=(c('font'), self.fs-1, 'italic'))
        accent = c('wine') if self.req else c('border')
        self.config(highlightbackground=accent, highlightthickness=2 if self.req else 1)
        self.clear_btn.pack_forget()

    def set_path(self, p):
        self.path_var.set(p)
        self.icon_lbl.config(text='✓', fg=c('ok'))
        self.status_lbl.config(text=f'  {Path(p).name}',
                               fg=c('ok'), font=(c('font'), self.fs-1, 'bold'))
        self.config(highlightbackground=c('ok'), highlightthickness=2)
        self.clear_btn.pack(side='right')
        LOGGER.info("Файл выбран: %s", p)

    def get(self): return self.path_var.get()
    def is_set(self):
        p = self.get().strip()
        return bool(p) and Path(p).exists()

# ── Поиск ─────────────────────────────────────────────────────────────────────
class SearchWindow(tk.Toplevel):
    def __init__(self, parent, settings, fs):
        super().__init__(parent)
        self.title('Поиск по истории')
        self.geometry('900x560'); self.configure(bg=c('bg'))
        self.settings = settings; self.fs = fs
        self._build(); self._load()

    def _build(self):
        fs = self.fs
        hdr = tk.Frame(self, bg=c('wine')); hdr.pack(fill='x')
        tk.Frame(hdr, bg=c('gold'), height=3).pack(fill='x')
        tk.Label(hdr, text='Поиск по истории анкет',
                 font=(c('font'), fs+4, 'bold'), bg=c('wine'), fg='white'
                 ).pack(padx=20, pady=14, anchor='w')

        sf = tk.Frame(self, bg=c('bg'), pady=12); sf.pack(fill='x', padx=20)
        tk.Label(sf, text='Введите ФИО, телефон, ИИН, дату или имя менеджера:',
                 font=(c('font'), fs-1, 'bold'), bg=c('bg'), fg=c('ink2')
                 ).pack(anchor='w', pady=(0,5))
        sr = tk.Frame(sf, bg=c('bg')); sr.pack(fill='x')
        self.q = tk.StringVar(); self.q.trace('w', lambda *_: self._filter())
        e = tk.Entry(sr, textvariable=self.q, font=(c('font'), fs+1), width=48,
                     relief='flat', bg=c('surface'), fg=c('ink'),
                     insertbackground=c('wine'),
                     highlightbackground=c('border_focus'), highlightthickness=2)
        e.pack(side='left', ipady=9, fill='x', expand=True, padx=(0,10)); e.focus()
        SecondaryButton(sr, '📊 Открыть базу', self._open_db, fs=fs-1).pack(side='left')

        frame = tk.Frame(self, bg=c('bg')); frame.pack(fill='both', expand=True, padx=20, pady=(0,8))
        style = ttk.Style()
        style.configure('S.Treeview', background=c('surface'), foreground=c('ink'),
                         fieldbackground=c('surface'), rowheight=26, font=(c('font'), fs-1))
        style.configure('S.Treeview.Heading', background=c('bg2'), foreground=c('ink2'),
                         font=(c('font'), fs-1, 'bold'))
        style.map('S.Treeview', background=[('selected', c('wine'))],
                  foreground=[('selected', 'white')])

        cols = ('phone','date','mgr','fio','iin','goal','file')
        hdrs = ('Телефон','Дата','Менеджер','ФИО клиента','ИИН','Цель','Файл')
        widths = (110,120,130,200,110,110,200)
        self.tree = ttk.Treeview(frame, columns=cols, show='headings',
                                  style='S.Treeview', height=16)
        for col, h, w in zip(cols, hdrs, widths):
            self.tree.heading(col, text=h); self.tree.column(col, width=w, anchor='w')
        sb  = ttk.Scrollbar(frame, orient='vertical',   command=self.tree.yview)
        sbx = ttk.Scrollbar(frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb.set, xscrollcommand=sbx.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        sb.grid(row=0, column=1, sticky='ns')
        sbx.grid(row=1, column=0, sticky='ew')
        frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)
        self.tree.bind('<Double-Button-1>', self._open)
        self.tree.tag_configure('ok',   foreground=c('ok'))
        self.tree.tag_configure('miss', foreground=c('ink3'))

        bot = tk.Frame(self, bg=c('bg'), pady=8); bot.pack(fill='x', padx=20)
        PrimaryButton(bot, '↗ Открыть файл', self._open, fs=fs-1,
                      padx=16, pady=8).pack(side='left', padx=(0,8))
        tk.Label(bot, text='Двойной клик — открыть файл',
                 font=(c('font'), fs-2, 'italic'),
                 bg=c('bg'), fg=c('ink3')).pack(side='left')
        SecondaryButton(bot, 'Закрыть', self.destroy, fs=fs-1,
                        padx=16, pady=8).pack(side='right')

    def _load(self):
        self.records = []
        for e in self.settings.get('recent_files', []):
            p = Path(e.get('path',''))
            self.records.append({'phone': e.get('phone',''), 'ts': e.get('ts',''),
                'mgr': e.get('manager_name',''), 'fio': e.get('fio',''),
                'iin': e.get('iin',''), 'goal': e.get('goal',''), 'path': p})
        known = {str(r['path']) for r in self.records}
        hist = Path('history')
        if hist.exists():
            for f in sorted(hist.rglob('*.xlsx'),
                            key=lambda x: x.stat().st_mtime, reverse=True):
                if str(f) not in known:
                    self.records.append({'phone':'',
                        'ts': datetime.fromtimestamp(f.stat().st_mtime).strftime('%d.%m.%Y %H:%M'),
                        'mgr':'', 'fio': f.stem, 'iin':'', 'goal':'', 'path': f})
        self._visible = self.records; self._show(self.records)

    def _show(self, recs):
        self.tree.delete(*self.tree.get_children()); self._visible = recs
        for r in recs:
            ok = r['path'].exists()
            self.tree.insert('','end', values=(
                r['phone'], r['ts'], r['mgr'], r['fio'], r['iin'], r['goal'],
                ('✓  ' if ok else '✗  ') + r['path'].name),
                tags=('ok' if ok else 'miss',))

    def _filter(self):
        q = self.q.get().strip().lower()
        if not q: self._show(self.records); return
        self._show([r for r in self.records if
            q in r['fio'].lower() or q in r['iin'].lower() or
            q in r['phone'].lower() or q in r['ts'] or
            q in r['mgr'].lower() or q in r['path'].name.lower()])

    def _open(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        p = self._visible[self.tree.index(sel[0])]['path']
        if p.exists(): os.startfile(str(p))
        else: messagebox.showwarning('Не найден', f'Файл удалён:\n{p}')

    def _open_db(self):
        if DB_FILE.exists(): os.startfile(str(DB_FILE))
        else: messagebox.showinfo('База', 'База анкет ещё не создана.')

# ── Статистика ────────────────────────────────────────────────────────────────
class StatsWindow(tk.Toplevel):
    def __init__(self, parent, settings, fs):
        super().__init__(parent)
        self.title('Статистика')
        self.geometry('680x560'); self.configure(bg=c('bg'))
        self.settings = settings; self.fs = fs
        self._build()

    def _build(self):
        fs = self.fs
        hdr = tk.Frame(self, bg=c('wine')); hdr.pack(fill='x')
        tk.Frame(hdr, bg=c('gold'), height=3).pack(fill='x')
        tk.Label(hdr, text='Статистика по анкетам',
                 font=(c('font'), fs+4, 'bold'), bg=c('wine'), fg='white'
                 ).pack(padx=20, pady=14, anchor='w')

        pf = Card(self); pf.pack(fill='x', padx=20, pady=12)
        pi = tk.Frame(pf, bg=c('surface')); pi.pack(padx=16, pady=10, anchor='w')
        tk.Label(pi, text='Период:', font=(c('font'), fs, 'bold'),
                 bg=c('surface'), fg=c('ink2')).pack(side='left', padx=(0,12))
        self.pv = tk.StringVar(value='7')
        for v, l in [('1','Сегодня'),('7','7 дней'),('30','30 дней'),('0','Всё время')]:
            tk.Radiobutton(pi, text=l, variable=self.pv, value=v,
                           font=(c('font'), fs), bg=c('surface'), fg=c('ink'),
                           selectcolor=c('wine_light'), activebackground=c('surface'),
                           activeforeground=c('wine'), command=self._refresh
                           ).pack(side='left', padx=8)

        tbl = tk.Frame(self, bg=c('bg')); tbl.pack(fill='x', padx=20)
        style = ttk.Style()
        style.configure('St.Treeview', background=c('surface'), foreground=c('ink'),
                         fieldbackground=c('surface'), rowheight=28, font=(c('font'), fs))
        style.configure('St.Treeview.Heading', background=c('bg2'), foreground=c('ink2'),
                         font=(c('font'), fs, 'bold'))
        style.map('St.Treeview', background=[('selected', c('wine'))],
                  foreground=[('selected', 'white')])
        cols = ('mgr','today','period','total')
        self.tree = ttk.Treeview(tbl, columns=cols, show='headings',
                                  style='St.Treeview', height=11)
        for col, h, w, a in zip(cols,['Менеджер','Сегодня','За период','Всего'],
                                 [240,100,110,100],['w','center','center','center']):
            self.tree.heading(col, text=h); self.tree.column(col, width=w, anchor=a)
        self.tree.pack(fill='x')
        self.tree.tag_configure('tot', font=(c('font'), fs, 'bold'), foreground=c('wine'))

        tk.Label(self, text='Активность за 14 дней:',
                 font=(c('font'), fs, 'bold'),
                 bg=c('bg'), fg=c('ink2')).pack(anchor='w', padx=20, pady=(14,4))
        self.days = tk.Text(self, height=7, font=('Consolas', fs),
                             bg='#2C1810', fg='#C4A882',
                             state='disabled', relief='flat', padx=14, pady=8)
        self.days.pack(fill='x', padx=20)
        SecondaryButton(self, '↺ Обновить', self._refresh,
                        fs=fs-1, padx=16, pady=8).pack(pady=12)
        self._refresh()

    def _refresh(self):
        recs = []; hist = Path('history')
        if hist.exists():
            for f in hist.rglob('*.xlsx'):
                try:
                    mt = datetime.fromtimestamp(f.stat().st_mtime)
                    mg = ''
                    parts = f.relative_to(hist).parts
                    if len(parts) > 2:
                        for m in self.settings.get('managers',[]):
                            if mgr_folder(m['name']) == parts[0]: mg = m['id']; break
                    recs.append((mt, mg))
                except: pass
        now = datetime.now(); today = now.date()
        dn  = int(self.pv.get())
        cut = now - timedelta(days=dn) if dn > 0 else datetime.min
        mm  = {m['id']: m['name'] for m in self.settings.get('managers',[])}
        cnt = defaultdict(lambda: {'today':0,'period':0,'total':0})
        for mt, mg in recs:
            cnt[mg]['total'] += 1
            if mt.date() == today: cnt[mg]['today'] += 1
            if mt >= cut:          cnt[mg]['period'] += 1
        self.tree.delete(*self.tree.get_children())
        tt=td=tp=0
        for mid, name in mm.items():
            v = cnt[mid]
            self.tree.insert('','end', values=(name,v['today'],v['period'],v['total']))
            tt+=v['total']; td+=v['today']; tp+=v['period']
        self.tree.insert('','end', values=('ИТОГО',td,tp,tt), tags=('tot',))
        dc = defaultdict(int)
        for mt, _ in recs:
            d = mt.date()
            if d >= today-timedelta(days=13): dc[d] += 1
        self.days.config(state='normal'); self.days.delete('1.0','end')
        mx = max(dc.values()) if dc else 1
        for i in range(13,-1,-1):
            d = today-timedelta(days=i)
            n = dc.get(d,0)
            lbl = 'Сегодня     ' if i==0 else d.strftime('%d.%m.%Y   ')
            bar = '█' * int(n/mx*30) if mx > 0 else ''
            self.days.insert('end', f'  {lbl}  {bar}{"  "+str(n) if n else ""}\n')
        self.days.config(state='disabled')

# ── Настройки ─────────────────────────────────────────────────────────────────
class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, settings, on_save):
        super().__init__(parent)
        self.title('Настройки')
        self.geometry('540x520'); self.resizable(False,False); self.grab_set()
        self.configure(bg=c('bg'))
        self.settings = settings; self.on_save = on_save
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=c('wine')); hdr.pack(fill='x')
        tk.Frame(hdr, bg=c('gold'), height=3).pack(fill='x')
        tk.Label(hdr, text='Настройки приложения',
                 font=(c('font'), 13, 'bold'), bg=c('wine'), fg='white'
                 ).pack(padx=16, pady=12, anchor='w')

        style = ttk.Style()
        style.configure('Set.TNotebook', background=c('bg'))
        style.configure('Set.TNotebook.Tab', background=c('surface'),
                         foreground=c('ink2'), padding=[14,7])
        style.map('Set.TNotebook.Tab',
                  background=[('selected', c('wine'))],
                  foreground=[('selected', 'white')])

        nb = ttk.Notebook(self, style='Set.TNotebook')
        nb.pack(fill='both', expand=True, padx=16, pady=12)

        # Менеджеры
        mt = tk.Frame(nb, bg=c('bg')); nb.add(mt, text='  Менеджеры  ')
        tk.Label(mt, text='Укажите настоящие имена. Они отображаются в списке\n'
                          'и используются для именования папок с анкетами.',
                 font=(c('font'), 9), fg=c('ink3'), bg=c('bg'), justify='left'
                 ).pack(anchor='w', padx=14, pady=10)
        canvas = tk.Canvas(mt, height=260, bg=c('bg'), highlightthickness=0)
        sb2 = ttk.Scrollbar(mt, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb2.set)
        canvas.pack(side='left', fill='both', expand=True, padx=14)
        sb2.pack(side='right', fill='y', pady=4)
        inner = tk.Frame(canvas, bg=c('bg'))
        canvas.create_window((0,0), window=inner, anchor='nw')
        inner.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        self.me = []
        for i, m in enumerate(self.settings['managers']):
            row = tk.Frame(inner, bg=c('bg')); row.pack(fill='x', pady=4)
            tk.Label(row, text=f'{i+1}.', width=3, bg=c('bg'),
                     font=(c('font'), 10), fg=c('ink3')).pack(side='left')
            var = tk.StringVar(value=m['name'])
            e = tk.Entry(row, textvariable=var, font=(c('font'), 10), width=34,
                          relief='flat', bg=c('surface'), fg=c('ink'),
                          insertbackground=c('wine'),
                          highlightbackground=c('border'),
                          highlightcolor=c('border_focus'), highlightthickness=1)
            e.pack(side='left', padx=6, ipady=6)
            self.me.append((m['id'], var))

        # Шрифт
        ft = tk.Frame(nb, bg=c('bg')); nb.add(ft, text='  Шрифт  ')
        tk.Label(ft, text='Размер шрифта:', font=(c('font'), 10, 'bold'),
                 fg=c('ink2'), bg=c('bg')).pack(anchor='w', padx=16, pady=(18,8))
        self.fsv = tk.IntVar(value=self.settings.get('font_size', 10))
        for sz, l in [(9,'Мелкий (9)'),(10,'Обычный (10)'),(12,'Крупный (12)'),(14,'Очень крупный (14)')]:
            tk.Radiobutton(ft, text=l, variable=self.fsv, value=sz,
                           font=(c('font'), 11), bg=c('bg'), fg=c('ink'),
                           selectcolor=c('wine_light'), activebackground=c('bg'),
                           activeforeground=c('wine')).pack(anchor='w', padx=32, pady=4)

        # CRM — автопередача лидов
        ct = tk.Frame(nb, bg=c('bg')); nb.add(ct, text='  CRM  ')
        tk.Label(ct, text='Автопередача лидов в веб-CRM после создания анкеты.\n'
                          'Адрес: https://crm-ipo.vercel.app\n'
                          'Ключ: значение PKB_API_KEY (или WEBHOOK_SECRET) из Vercel.\n'
                          'Оставьте пустым, чтобы отключить.',
                 font=(c('font'), 9), fg=c('ink3'), bg=c('bg'), justify='left'
                 ).pack(anchor='w', padx=16, pady=14)
        tk.Label(ct, text='Адрес CRM:', font=(c('font'), 9, 'bold'),
                 fg=c('ink2'), bg=c('bg')).pack(anchor='w', padx=16)
        self.crm_url_v = tk.StringVar(value=self.settings.get('crm_url', ''))
        tk.Entry(ct, textvariable=self.crm_url_v, font=(c('font'), 10), width=44,
                 relief='flat', bg=c('surface'), fg=c('ink'), insertbackground=c('wine'),
                 highlightbackground=c('border'), highlightthickness=1
                 ).pack(anchor='w', padx=16, pady=(3, 10), ipady=6)
        tk.Label(ct, text='API-ключ:', font=(c('font'), 9, 'bold'),
                 fg=c('ink2'), bg=c('bg')).pack(anchor='w', padx=16)
        self.crm_key_v = tk.StringVar(value=self.settings.get('crm_key', ''))
        tk.Entry(ct, textvariable=self.crm_key_v, font=(c('font'), 10), width=44, show='•',
                 relief='flat', bg=c('surface'), fg=c('ink'), insertbackground=c('wine'),
                 highlightbackground=c('border'), highlightthickness=1
                 ).pack(anchor='w', padx=16, pady=(3, 10), ipady=6)

        # Бэкап
        bt = tk.Frame(nb, bg=c('bg')); nb.add(bt, text='  Бэкап  ')
        tk.Label(bt, text='Папка для автоматического бэкапа анкет.\nОставьте пустым, чтобы отключить.',
                 font=(c('font'), 9), fg=c('ink3'), bg=c('bg'), justify='left'
                 ).pack(anchor='w', padx=16, pady=14)
        br = tk.Frame(bt, bg=c('bg')); br.pack(fill='x', padx=16)
        self.bkv = tk.StringVar(value=self.settings.get('backup_folder',''))
        e = tk.Entry(br, textvariable=self.bkv, font=(c('font'), 10), width=36,
                      relief='flat', bg=c('surface'), fg=c('ink'),
                      insertbackground=c('wine'),
                      highlightbackground=c('border'), highlightthickness=1)
        e.pack(side='left', ipady=7)
        SecondaryButton(br, '📁',
                        lambda: self.bkv.set(filedialog.askdirectory() or self.bkv.get()),
                        fs=10, padx=10, pady=7).pack(side='left', padx=8)

        bf = tk.Frame(self, bg=c('bg'), pady=12); bf.pack(fill='x', padx=16)
        PrimaryButton(bf, '  Сохранить  ', self._save,
                      fs=10, padx=20, pady=10).pack(side='left', padx=(0,10))
        SecondaryButton(bf, 'Отмена', self.destroy,
                        fs=10, padx=20, pady=10).pack(side='left')

    def _save(self):
        for mid, var in self.me:
            for m in self.settings['managers']:
                if m['id'] == mid: m['name'] = var.get().strip() or mid
        self.settings['font_size']     = self.fsv.get()
        self.settings['backup_folder'] = self.bkv.get().strip()
        self.settings['crm_url']       = self.crm_url_v.get().strip()
        self.settings['crm_key']       = self.crm_key_v.get().strip()
        save_settings(self.settings); self.on_save(); self.destroy()

# ── Главное окно ──────────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        self.settings = load_settings()
        self.last_out = None
        self._build()

    def _build(self):
        fs = self.settings.get('font_size', 10)
        self.fs = fs
        self.root.title('ПКБ — Анкета клиента')
        self.root.geometry('960x920')
        self.root.minsize(860, 600)
        self.root.configure(bg=c('bg'))

        style = ttk.Style(); style.theme_use('clam')
        style.configure('TProgressbar', troughcolor=c('border'),
                         background=c('wine'), thickness=4)
        style.configure('TCombobox', fieldbackground=c('surface_warm'),
                         background=c('surface'), foreground=c('ink'),
                         arrowcolor=c('ink2'))

        # Шапка — фиксированная, не скроллится
        self._build_header(fs)

        # Скроллируемая область для всего остального
        scroll_wrap = tk.Frame(self.root, bg=c('bg'))
        scroll_wrap.pack(fill='both', expand=True)

        self._canvas = tk.Canvas(scroll_wrap, bg=c('bg'),
                                  highlightthickness=0, bd=0)
        self._scrollbar = ttk.Scrollbar(scroll_wrap, orient='vertical',
                                         command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._scrollbar.set)

        self._scrollbar.pack(side='right', fill='y')
        self._canvas.pack(side='left', fill='both', expand=True)

        # Внутренний фрейм — контент
        self._inner = tk.Frame(self._canvas, bg=c('bg'))
        self._inner_id = self._canvas.create_window(
            (0, 0), window=self._inner, anchor='nw')

        # Обновляем размер scrollregion при изменении контента
        self._inner.bind('<Configure>', self._on_inner_configure)
        self._canvas.bind('<Configure>', self._on_canvas_configure)

        # Прокрутка колёсиком мыши
        self._canvas.bind_all('<MouseWheel>',
            lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))

        # Все секции рисуются внутри self._inner
        self._build_manager(fs)
        self._build_tip(fs)
        self._build_uploads(fs)
        self._build_extras(fs)
        self._build_actions(fs)
        self._build_status(fs)
        self._build_history(fs)
        self._build_log(fs)

    def _on_inner_configure(self, event):
        self._canvas.configure(scrollregion=self._canvas.bbox('all'))

    def _on_canvas_configure(self, event):
        # Растягиваем inner frame по ширине canvas
        self._canvas.itemconfig(self._inner_id, width=event.width)

    def _build_header(self, fs):
        wrap = tk.Frame(self.root, bg=c('wine')); wrap.pack(fill='x')
        tk.Frame(wrap, bg=c('gold'), height=4).pack(fill='x')
        inner = tk.Frame(wrap, bg=c('wine')); inner.pack(fill='x', padx=20, pady=14)
        logo = tk.Frame(inner, bg=c('wine')); logo.pack(side='left')
        tk.Label(logo, text='ПКБ', font=(c('font'), fs+8, 'bold'),
                 bg=c('wine'), fg=c('gold')).pack(side='left')
        tk.Label(logo, text='  Анкета клиента', font=(c('font'), fs+3),
                 bg=c('wine'), fg='#E8D5D8').pack(side='left', pady=6)
        nav = tk.Frame(inner, bg=c('wine')); nav.pack(side='right')
        badge_txt = self._sub_badge_text()
        if badge_txt:
            tk.Label(nav, text=badge_txt, font=(c('font'), fs-1, 'bold'),
                     bg=c('wine_dark'), fg=c('gold'), padx=12, pady=6
                     ).pack(side='left', padx=(0,12))
        for txt, cmd in [('Поиск', self.open_search),
                          ('Статистика', self.open_stats),
                          ('Настройки', self.open_settings)]:
            tk.Button(nav, text=txt, font=(c('font'), fs-1),
                      bg=c('wine'), fg='#E8D5D8',
                      activebackground=c('wine_dark'), activeforeground='white',
                      relief='flat', cursor='hand2',
                      padx=14, pady=6, command=cmd).pack(side='left', padx=3)
        tk.Frame(wrap, bg='#4E1C28', height=1).pack(fill='x')

    def _sub_badge_text(self):
        """Текст бейджа подписки в шапке. None — если подписка не настроена."""
        if not self.settings.get('license_ok_once'):
            return None
        until = self.settings.get('license_until')
        if not until:
            return 'Подписка активна'
        try:
            from license_gate import _parse_date
            d = _parse_date(until)
        except Exception:
            d = None
        if not d:
            return 'Подписка активна'
        from datetime import date
        days = (d - date.today()).days
        if days < 0:
            return 'Подписка истекла'
        return 'Подписка: {} дн.'.format(days)

    def _build_manager(self, fs):
        card = Card(self._inner); card.pack(fill='x', padx=16, pady=(12,0))
        inner = tk.Frame(card, bg=c('surface')); inner.pack(fill='x', padx=18, pady=12)
        SectionLabel(inner, 'Менеджер').pack(anchor='w', pady=(0,5))
        row = tk.Frame(inner, bg=c('surface')); row.pack(anchor='w')
        names = [m['name'] for m in self.settings['managers']]
        self.mgr_var = tk.StringVar(value=names[0])
        ttk.Combobox(row, textvariable=self.mgr_var, values=names,
                     width=26, state='readonly', font=(c('font'), fs+1)
                     ).pack(side='left')
        tk.Label(row, text='  ← Выберите своё имя перед началом',
                 font=(c('font'), fs-1, 'italic'),
                 fg=c('ink3'), bg=c('surface')).pack(side='left')

    def _build_tip(self, fs):
        tip = tk.Frame(self._inner, bg=c('gold_light'),
                        highlightbackground=c('gold'), highlightthickness=1)
        tip.pack(fill='x', padx=16, pady=8)
        tk.Label(tip,
                 text='Как работать:   '
                      '① Выберите менеджера   →   '
                      '② Загрузите кредитный отчёт   →   '
                      '③ Загрузите пенсионку (если есть)   →   '
                      '④ Заполните доп. поля   →   '
                      '⑤ Нажмите «Создать анкету»',
                 font=(c('font'), fs-1), bg=c('gold_light'), fg=c('warn'),
                 wraplength=900, justify='left').pack(padx=14, pady=10, anchor='w')

    def _build_uploads(self, fs):
        self.cz = UploadZone(self._inner,
            title='Кредитный отчёт из ПКБ',
            subtitle='Скачайте на 1cb.kz → раздел «Кредитный отчёт»',
            required=True, fs=fs)
        self.cz.pack(fill='x', padx=16, pady=(0,6))
        self.pz = UploadZone(self._inner,
            title='Пенсионная выписка',
            subtitle='Скачайте на enpf.kz → «Выписка об отчислениях» (не обязательно)',
            required=False, fs=fs)
        self.pz.pack(fill='x', padx=16)

    def _build_extras(self, fs):
        card = Card(self._inner); card.pack(fill='x', padx=16, pady=(8,0))
        tk.Frame(card, bg=c('gold'), height=3).pack(fill='x')
        inner = tk.Frame(card, bg=c('surface')); inner.pack(fill='x', padx=18, pady=(10,14))
        SectionLabel(inner, 'Дополнительная информация о клиенте').pack(anchor='w', pady=(0,10))

        row1 = tk.Frame(inner, bg=c('surface')); row1.pack(fill='x')
        self.f_phone  = InputField(row1, 'Номер телефона',    '+7 ___ ___ __ __',         18, fs)
        self.f_goal   = InputField(row1, 'Цель кредита',      'ипотека / авто / ремонт…', 26, fs)
        self.f_onhand = InputField(row1, 'Наличные на руках', 'сумма на руках клиента',   18, fs)
        for f in [self.f_phone, self.f_goal, self.f_onhand]:
            f.pack(side='left', padx=(0,20))

        tk.Frame(inner, bg=c('border'), height=1).pack(fill='x', pady=(12,0))

        dep_hdr = tk.Frame(inner, bg=c('surface')); dep_hdr.pack(fill='x', pady=(8,8))
        SectionLabel(dep_hdr, 'Депозитные накопления').pack(side='left')
        tk.Label(dep_hdr,
                 text='  Н — наличные на депозите · БВ — Баспана Вай · ГП — Госпремия',
                 font=(c('font'), fs-2, 'italic'),
                 bg=c('surface'), fg=c('ink3')).pack(side='left')

        row2 = tk.Frame(inner, bg=c('surface')); row2.pack(fill='x')
        self.f_cash = InputField(row2, 'Наличные (Н)',     'сумма', 14, fs)
        self.f_bv   = InputField(row2, 'Баспана Вай (БВ)', 'сумма', 14, fs)
        self.f_gp   = InputField(row2, 'Госпремия (ГП)',   'сумма', 14, fs)
        for f in [self.f_cash, self.f_bv, self.f_gp]:
            f.pack(side='left', padx=(0,16))

    def _build_actions(self, fs):
        card = Card(self._inner); card.pack(fill='x', padx=16, pady=8)
        inner = tk.Frame(card, bg=c('surface')); inner.pack(fill='x', padx=18, pady=14)
        self.gen_btn = PrimaryButton(inner, '▶  Создать анкету', self.start_gen,
                                      fs=fs+2, padx=28, pady=14)
        self.gen_btn.pack(side='left', padx=(0,10))
        self.open_btn = SecondaryButton(inner, '↗  Открыть последний файл',
                                         self.open_last, fs=fs-1, padx=16, pady=14)
        self.open_btn.config(state='disabled'); self.open_btn.pack(side='left', padx=(0,8))
        SecondaryButton(inner, '📥  Импорт анкет', self.import_anketa,
                        fs=fs-1, padx=14, pady=14).pack(side='left', padx=(0,8))
        SecondaryButton(inner, '📋  База', self.open_db,
                        fs=fs-1, padx=14, pady=14).pack(side='left', padx=(0,8))
        SecondaryButton(inner, '📂  Все анкеты', self.open_hist,
                        fs=fs-1, padx=14, pady=14).pack(side='left')

    def _build_status(self, fs):
        self.progress = ttk.Progressbar(self._inner, mode='indeterminate',
                                         style='TProgressbar')
        self.progress.pack(fill='x', padx=16, pady=(0,4))
        self.sv = tk.StringVar(value='Готов к работе')
        self.sl = tk.Label(self._inner, textvariable=self.sv,
                            font=(c('font'), fs), bg=c('surface'), fg=c('ok'),
                            anchor='w', pady=8, padx=16,
                            highlightbackground=c('border'), highlightthickness=1)
        self.sl.pack(fill='x', padx=16, pady=(0,4))

    def _build_history(self, fs):
        card = Card(self._inner); card.pack(fill='x', padx=16, pady=(0,4))
        inner = tk.Frame(card, bg=c('surface')); inner.pack(fill='x', padx=18, pady=(10,0))
        hr = tk.Frame(inner, bg=c('surface')); hr.pack(fill='x', pady=(0,6))
        SectionLabel(hr, 'Последние анкеты').pack(side='left')
        tk.Label(hr, text='  двойной клик — открыть файл',
                 font=(c('font'), fs-2, 'italic'),
                 fg=c('ink3'), bg=c('surface')).pack(side='left')
        self.hl = tk.Listbox(card, height=4, font=('Consolas', fs-1),
                              bg=c('surface'), fg=c('ink2'),
                              selectbackground=c('wine'), selectforeground='white',
                              activestyle='none', relief='flat',
                              cursor='hand2', bd=0, highlightthickness=0)
        self.hl.pack(fill='x', padx=18, pady=(0,10))
        self.hl.bind('<Double-Button-1>', self._open_hist_item)
        self._refresh_hist()

    def _build_log(self, fs):
        card = Card(self._inner); card.pack(fill='both', expand=True, padx=16, pady=(0,12))
        inner = tk.Frame(card, bg=c('surface')); inner.pack(fill='x', padx=18, pady=(10,4))
        SectionLabel(inner, 'Лог выполнения').pack(anchor='w')
        self.log = scrolledtext.ScrolledText(card, height=5, state='disabled',
                                              font=('Consolas', 8),
                                              bg='#2C1810', fg='#C4A882',
                                              relief='flat', bd=0,
                                              padx=10, pady=6, highlightthickness=0)
        self.log.pack(fill='both', expand=True, padx=18, pady=(0,10))
        self.log.tag_config('err', foreground='#F87171')
        self.log.tag_config('wrn', foreground='#FCD34D')
        self.log.tag_config('inf', foreground='#C4A882')
        LOGGER.handlers.clear()
        LOGGER.addHandler(TextHandler(self.log))
        fh = logging.FileHandler('app.log', encoding='utf-8', mode='a')
        fh.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s'))
        LOGGER.addHandler(fh)
        LOGGER.addHandler(logging.StreamHandler(sys.stdout))
        LOGGER.info("=== ПКБ Анкета v18 ===")

    def _refresh_hist(self):
        self.hl.delete(0,'end')
        for e in self.settings.get('recent_files',[])[:8]:
            p  = Path(e.get('path',''))
            ok = '✓' if p.exists() else '✗'
            ph = f"[{e['phone']}]  " if e.get('phone') else ''
            self.hl.insert('end', f"  {ok}  {e.get('ts','')}   {ph}{e.get('fio','')}   —   {p.name}")

    def _add_hist(self, path, fio, extra):
        entry = {'path':str(path), 'ts':datetime.now().strftime('%d.%m.%Y %H:%M'),
                 'fio':fio, **extra}
        recent = [r for r in self.settings.get('recent_files',[])
                  if r.get('path')!=str(path)]
        recent.insert(0, entry); self.settings['recent_files'] = recent[:30]
        save_settings(self.settings); self._refresh_hist()

    def _open_hist_item(self, _=None):
        sel = self.hl.curselection()
        if not sel: return
        recent = self.settings.get('recent_files',[])
        if sel[0] < len(recent):
            p = Path(recent[sel[0]]['path'])
            if p.exists(): os.startfile(str(p))
            else: messagebox.showwarning('Не найден', f'Файл удалён:\n{p}')

    def open_search(self):   SearchWindow(self.root, self.settings, self.fs)
    def open_stats(self):    StatsWindow(self.root, self.settings, self.fs)
    def open_settings(self):
        def on_save():
            self.settings = load_settings()
            for w in self.root.winfo_children(): w.destroy()
            self._build()
        SettingsDialog(self.root, self.settings, on_save)

    def open_last(self):
        if self.last_out and Path(self.last_out).exists():
            os.startfile(str(self.last_out))
        else: messagebox.showinfo('Нет файла', 'Последний файл не найден.')

    def open_db(self):
        if DB_FILE.exists(): os.startfile(str(DB_FILE))
        else: messagebox.showinfo('База пуста', 'Пока ни одной анкеты не создано.')

    def open_hist(self):
        h = Path('history').absolute()
        if h.exists(): os.startfile(str(h))

    def import_anketa(self):
        """Импорт готовых анкет (.xlsx) → строки в база_анкет.xlsx."""
        paths = filedialog.askopenfilenames(
            title='Выберите анкеты для импорта в базу',
            filetypes=[('Excel анкеты', '*.xlsx'), ('Все файлы', '*.*')])
        if not paths:
            return
        mgr_name = self._get_mgr()[1]
        ok = 0; fail = 0; empty = 0
        for p in paths:
            try:
                rec = read_anketa_xlsx(p)
                if not (rec.get('fio') or rec.get('iin') or rec.get('phone')):
                    empty += 1
                    LOGGER.warning("Импорт: пусто — %s", Path(p).name)
                    continue
                if not rec.get('manager'):
                    rec['manager'] = mgr_name
                update_db(rec)
                post_to_crm(self.settings, rec)  # автопередача в CRM (если настроено)
                ok += 1
                LOGGER.info("Импорт: %s (ФИО=%s ИИН=%s)", Path(p).name, rec.get('fio'), rec.get('iin'))
            except Exception as e:
                fail += 1
                LOGGER.warning("Импорт: ошибка %s — %s", Path(p).name, e)
        msg = f'Импортировано в базу: {ok}'
        if empty: msg += f'\nПустых (пропущено): {empty}'
        if fail:  msg += f'\nОшибок: {fail}'
        messagebox.showinfo('Импорт анкет', msg)
        self._status(f'📥 Импорт: добавлено {ok} строк в базу', c('ok'))

    def _get_mgr(self):
        name = self.mgr_var.get()
        for m in self.settings['managers']:
            if m['name'] == name: return m['id'], m['name']
        return 'manager1', name

    def start_gen(self):
        if not self.cz.is_set():
            messagebox.showwarning('Файл не выбран',
                'Выберите кредитный отчёт из ПКБ.\n\n'
                'Нажмите «Выбрать PDF» в блоке «Кредитный отчёт».')
            return
        self.gen_btn.config(state='disabled', bg='#A0948A', activebackground='#A0948A')
        self.open_btn.config(state='disabled')
        self.progress.start()
        self._status('Обрабатываю документы…', c('warn'))
        Thread(target=self._gen, daemon=True).start()

    def _gen(self):
        try:
            credit_pdf = Path(self.cz.get())
            pen_str    = self.pz.get().strip()
            pen_pdf    = Path(pen_str) if pen_str else None
            mgr_id, mgr_name = self._get_mgr()
            phone  = self.f_phone.get()
            goal   = self.f_goal.get()
            onhand = self.f_onhand.get()
            cash   = self.f_cash.get()
            bv     = self.f_bv.get()
            gp     = self.f_gp.get()

            tmpl = Path('assets/template.xlsx')
            if not tmpl.exists():
                raise FileNotFoundError('Шаблон не найден: assets\\template.xlsx')

            self._status('Шаг 1 из 3  ·  читаю кредитный отчёт…', c('warn'))
            cd = extract_any_credit_report(pdf_path=credit_pdf, enable_ocr=False,
                    ocr_lang='rus', poppler_path=None,
                    tesseract_cmd=None, tessdata_dir=None, plumber_layout=False)
            validate_credit_data(cd, strict=False)

            # Если ничего не прочиталось — показываем понятный диалог
            if not cd.get('iin') and not cd.get('pkr') and not cd.get('active'):
                raise ValueError(
                    'Не удалось прочитать кредитный отчёт.\n\n'
                    'Возможные причины:\n'
                    '• Файл не является отчётом ПКБ (1cb.kz) или ГКБ/МКБ (mkb.kz)\n'
                    '• PDF отсканирован (не текстовый)\n'
                    '• Формат отчёта изменился\n\n'
                    'Попробуйте скачать отчёт заново на 1cb.kz или id.mkb.kz'
                )

            LOGGER.info("%s: ПКР=%s ИИН=%s активных=%s завершённых=%s старых=%s отозванных=%s",
                (cd.get('source') or 'ПКБ').upper(),
                cd.get('pkr'), mask_iin(cd.get('iin')),
                len(cd.get('active') or []),
                len(cd.get('completed_recent') or []),
                len(cd.get('completed_old') or []),
                len(cd.get('revoked') or []))

            monthly = sum((x.periodic_payment or 0) for x in (cd.get('active') or []))

            pd = None; srzp = 0
            if pen_pdf and pen_pdf.exists():
                self._status('Шаг 2 из 3  ·  читаю пенсионную выписку…', c('warn'))
                pd = parse_pension_pdf_tables(pen_pdf)
                validate_pension_data(pd, strict=False)
                srzp = calc_pension_avg(pd.get('rows') or []).get('avg_salary', 0)
                LOGGER.info("Пенсионка: ФИО=%s СРЗП=%s", pd.get('fio'), srzp)
            else:
                LOGGER.info("Пенсионка не выбрана")

            fio      = (pd.get('fio') if pd else None) or cd.get('fio') or (cd.get('iin') or 'UNKNOWN')
            iin      = cd.get('iin') or (pd.get('iin') if pd else '')
            safe_fio = safe_fn(fio)
            safe_iin = safe_fn(mask_iin(iin))
            ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
            ddir     = datetime.now().strftime('%Y-%m-%d')
            mfold    = mgr_folder(mgr_name)
            out_dir  = Path('history') / mfold / ddir
            out_dir.mkdir(parents=True, exist_ok=True)

            existing = list(out_dir.glob(f'{mgr_id}_{safe_iin}_*.xlsx'))
            if existing:
                newest  = max(existing, key=lambda p: p.stat().st_mtime)
                age_min = (datetime.now().timestamp() - newest.stat().st_mtime) / 60
                if age_min < 60:
                    ans = messagebox.askyesno('Анкета уже существует',
                        f'Анкета на этого клиента создана {int(age_min)} мин. назад:\n'
                        f'{newest.name}\n\nСоздать новую копию?')
                    if not ans:
                        self._status('Готов к работе', c('ok'))
                        self.root.after(0, self._finish); return

            ph_pfx   = f"{safe_fn(phone)}_" if phone else ""
            fname    = f"{mgr_id}_{ph_pfx}{safe_iin}_{safe_fio}_{ts}.xlsx"
            out_path = out_dir / fname

            self._status('Шаг 3 из 3  ·  формирую Excel-анкету…', c('warn'))
            write_output_legacy_excel(
                template_path=tmpl, output_path=out_path,
                pkr=cd.get('pkr'),
                active=cd.get('active') or [],
                recent=cd.get('completed_recent') or [],
                old=cd.get('completed_old') or [],
                revoked=cd.get('revoked') or [],
                blank_zero=True, only_loans_active=False)

            self._write_extra(out_path, phone, goal, onhand, cash, bv, gp)

            if pd:
                fill_anketa_from_pension(out_path, pd, sort_rows=False)

            bk = self.settings.get('backup_folder','').strip()
            if bk:
                try:
                    bd = Path(bk) / mfold / ddir
                    bd.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(str(out_path), str(bd/fname))
                    LOGGER.info("Бэкап: %s", bd/fname)
                except Exception as be:
                    LOGGER.warning("Бэкап: %s", be)

            rec_db = {
                'phone': phone, 'date': datetime.now().strftime('%d.%m.%Y %H:%M'),
                'manager': mgr_name, 'fio': fio, 'iin': iin,
                'pkr': str(cd.get('pkr') or ''),
                'load': str(monthly) if monthly else '',
                'srzp': str(srzp) if srzp else '',
                'goal': goal, 'onhand': onhand,
                'cash': cash, 'bv': bv, 'gp': gp,
                'path': str(out_path),
            }
            update_db(rec_db)
            # Автопередача лида в CRM (если настроено) — вместе с кредитной историей
            post_to_crm(self.settings, rec_db, contracts_to_credits(cd.get('active')))

            self.last_out = str(out_path)
            self._add_hist(out_path, fio, {
                'phone': phone, 'manager_name': mgr_name,
                'iin': iin, 'pkr': str(cd.get('pkr') or ''),
                'srzp': str(srzp), 'goal': goal,
            })
            LOGGER.info("Готово: %s", out_path)
            self.root.after(0, lambda: self._success(fname, out_path, out_dir))

        except Exception as e:
            LOGGER.exception("Ошибка")
            self.root.after(0, lambda: self._error(str(e)))
        finally:
            self.root.after(0, self._finish)

    def _write_extra(self, path, phone, goal, onhand, cash, bv, gp):
        try:
            import openpyxl
            from openpyxl.cell.cell import MergedCell
            from openpyxl.utils import coordinate_to_tuple
            wb = openpyxl.load_workbook(str(path))
            if 'Анкета' not in wb.sheetnames:
                wb.save(str(path)); return
            ws = wb['Анкета']
            def ss(ref, val):
                if not val: return
                cell = ws[ref]
                if isinstance(cell, MergedCell):
                    r, col = coordinate_to_tuple(ref)
                    for mc in ws.merged_cells.ranges:
                        if mc.min_row<=r<=mc.max_row and mc.min_col<=col<=mc.max_col:
                            ws.cell(mc.min_row, mc.min_col).value = val; return
                else: cell.value = val
            if phone:  ss('F6', phone)
            # Цель: строка 8 (Брак), крайнее правое поле J8 — свободно
            if goal:   ss('H10', goal)
            if onhand:
                ss('G13', 'Нал. на руках:')
                ss('H13', onhand)
            if cash:   ss('E18', cash)
            if bv:     ss('H18', bv)
            if gp:     ss('J18', gp)
            wb.save(str(path))
        except Exception as e:
            LOGGER.warning("Доп. поля: %s", e)

    def _success(self, fname, out_path, out_dir):
        self._status(f'✓  Анкета создана:  {fname}', c('ok'))
        self.open_btn.config(state='normal')
        win = tk.Toplevel(self.root)
        win.title('Анкета создана')
        win.geometry('500x250'); win.resizable(False,False); win.grab_set()
        win.configure(bg=c('surface'))
        tk.Frame(win, bg=c('ok'), height=4).pack(fill='x')
        body = tk.Frame(win, bg=c('surface')); body.pack(fill='both', expand=True, padx=28, pady=20)
        tk.Label(body, text='✓  Анкета успешно создана',
                 font=(c('font'), 15, 'bold'), bg=c('surface'), fg=c('ok')).pack(anchor='w')
        tk.Label(body, text=fname, font=(c('font'), 9),
                 bg=c('surface'), fg=c('ink3'), wraplength=440, justify='left'
                 ).pack(anchor='w', pady=(6,0))
        tk.Frame(body, bg=c('border'), height=1).pack(fill='x', pady=16)
        bf = tk.Frame(body, bg=c('surface')); bf.pack(anchor='w')
        PrimaryButton(bf, '↗  Открыть файл',
                      lambda: [os.startfile(str(out_path)), win.destroy()],
                      fs=9, padx=14, pady=9).pack(side='left', padx=(0,8))
        for txt, cmd in [
            ('📁  Папка',   lambda: [os.startfile(str(out_dir)), win.destroy()]),
            ('🖨  Печать',  lambda: [self._print(out_path), win.destroy()]),
            ('📋  База',    lambda: [self.open_db(), win.destroy()]),
        ]:
            SecondaryButton(bf, txt, cmd, fs=9, padx=12, pady=9).pack(side='left', padx=(0,6))
        SecondaryButton(body, 'Закрыть', win.destroy,
                        fs=9, padx=12, pady=8).pack(anchor='e', pady=(10,0))

    def _print(self, path):
        try: os.startfile(str(path), 'print')
        except: os.startfile(str(path))

    def _error(self, msg):
        self._status('✗  Ошибка — смотри лог внизу', c('err'))
        messagebox.showerror('Ошибка при создании анкеты',
            f'{msg}\n\nПодробности в разделе «Лог» внизу окна.')

    def _status(self, text, color=None):
        def _():
            self.sv.set(text)
            if color: self.sl.config(fg=color)
        self.root.after(0, _)

    def _finish(self):
        self.progress.stop()
        self.gen_btn.config(state='normal', bg=c('wine'), activebackground=c('wine_dark'))


def main():
    # TkinterDnD.Tk() включает перетаскивание PDF в окно (drag-and-drop).
    # Если библиотека не установлена — обычное окно, работает кнопка «Выбрать».
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except Exception:
        root = tk.Tk()
    root.withdraw()  # прячем главное окно до проверки подписки

    # Лицензионный гейт (подписка «по времени»). Если license_url пуст — пропускается.
    try:
        from license_gate import require_license
        settings = load_settings()
        if not require_license(root, settings, save_settings, logger=LOGGER):
            root.destroy()
            return
    except Exception as e:
        LOGGER.warning("Лицензия: пропущена из-за ошибки (%s)", e)

    root.deiconify()
    App(root)
    root.mainloop()

if __name__ == '__main__':
    main()
