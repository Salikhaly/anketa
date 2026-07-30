# -*- coding: utf-8 -*-
"""
main.py — PDF -> Excel extractor (Windows / PowerShell friendly)
v11.1: 2-sheet layout (Анкета + Кредиты), compact pension, per-org СРЗП.
FIXED: MergedCell write error — only write to top-left cells.

- Credit report PDF -> writes to legacy sheet "Кредиты"
- Pension PDF -> fills compact "Анкета": FIO, IIN, org tables (3 cols × 8 rows),
  per-organisation СРЗП
- Formula: СРЗП = (sum*7.9/6 + (sum-min-max)*7.9/(n-2)) / 2
- Robust paths: Windows + file: URIs
- Extraction: pdfplumber -> pypdf -> OCR fallback
"""

from __future__ import annotations

import argparse
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse, unquote

LOGGER = logging.getLogger("pdf_extract")


# ----------------------------
# Logging
# ----------------------------

def setup_logging(debug: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if debug else logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )


# ----------------------------
# PATHS: Windows + file: URI
# ----------------------------

def resolve_input_path(path_or_uri: str) -> Path:
    s = (path_or_uri or "").strip()
    if not s:
        raise ValueError("Пустой путь")
    if re.match(r"^[a-zA-Z]:[\\/]", s) or s.startswith("\\"):
        return Path(s)
    if s.lower().startswith("file:"):
        p = urlparse(s)
        combined = unquote((p.netloc or "") + (p.path or ""))
        if re.match(r"^/[a-zA-Z]:/", combined):
            combined = combined[1:]
        return Path(combined)
    return Path(s)


# ----------------------------
# Text & numbers
# ----------------------------

def normalize_ws(s: str) -> str:
    if s is None:
        return ""
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\s+\n", "\n", s)
    s = re.sub(r"\n\s+", "\n", s)
    return s.strip()


def extract_first(text: str, pat: str, flags: int = re.IGNORECASE | re.DOTALL) -> Optional[str]:
    m = re.search(pat, text or "", flags=flags)
    return normalize_ws(m.group(1)) if m else None


def parse_decimal_number(num_str: str) -> Decimal:
    s = (num_str or "").replace("\u00a0", " ").replace(" ", "").replace("'", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.]", "", s)
    if s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]
    if not s:
        raise ValueError(f"Не удалось распарсить число: {num_str!r}")
    return Decimal(s)


def money_to_int_kzt(num_str: str) -> int:
    return int(parse_decimal_number(num_str))


def parse_int_with_spaces(num_str: str) -> int:
    s = (num_str or "").replace("\u00a0", " ").replace(" ", "")
    s = re.sub(r"[^0-9]", "", s)
    if not s:
        raise ValueError(f"Не удалось распарсить int: {num_str!r}")
    return int(s)


def to_cell_str(val: Any, blank_zero: bool = True) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        if blank_zero and float(val) == 0.0:
            return None
        return str(int(val))
    if isinstance(val, str):
        if blank_zero and val.strip() in {"0", "0.0", "0,0"}:
            return None
        return val
    return str(val)


# ----------------------------
# PDF extraction
# ----------------------------

def extract_pages_text_pdfplumber(pdf_path: Path, layout: bool = False) -> List[str]:
    import pdfplumber
    out: List[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            t = page.extract_text(layout=layout) or ""
            out.append(t)
            LOGGER.debug("pdfplumber page=%s text_len=%s", i + 1, len(t.strip()))
    return out


def extract_pages_text_pypdf(pdf_path: Path) -> List[str]:
    from pypdf import PdfReader
    reader = PdfReader(str(pdf_path))
    out: List[str] = []
    for i, page in enumerate(reader.pages):
        t = page.extract_text(extraction_mode="layout") or ""
        out.append(t)
        LOGGER.debug("pypdf page=%s text_len=%s", i + 1, len(t.strip()))
    return out


def ocr_pages(
    pdf_path: Path,
    lang: str = "rus",
    dpi: int = 300,
    poppler_path: Optional[str] = None,
    tesseract_cmd: Optional[str] = None,
    tessdata_dir: Optional[str] = None,
) -> List[str]:
    from pdf2image import convert_from_path
    import pytesseract
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    config = f'--tessdata-dir "{tessdata_dir}"' if tessdata_dir else ""
    images = convert_from_path(str(pdf_path), dpi=dpi, poppler_path=poppler_path)
    out: List[str] = []
    for i, img in enumerate(images):
        txt = pytesseract.image_to_string(img, lang=lang, config=config) or ""
        out.append(txt)
        LOGGER.debug("ocr page=%s text_len=%s", i + 1, len(txt.strip()))
    return out


def is_probably_scanned(pages_text: List[str], sample_pages: int = 2, min_chars: int = 80) -> bool:
    n = min(sample_pages, len(pages_text))
    total = sum(len((pages_text[i] or "").strip()) for i in range(n))
    return total < min_chars


def extract_pages_text(
    pdf_path: Path,
    enable_ocr: bool,
    ocr_lang: str,
    poppler_path: Optional[str],
    tesseract_cmd: Optional[str],
    tessdata_dir: Optional[str],
    plumber_layout: bool = False,
) -> List[str]:
    pages: List[str] = []
    try:
        pages = extract_pages_text_pdfplumber(pdf_path, layout=plumber_layout)
    except Exception as e:
        LOGGER.warning("pdfplumber failed: %s", e)
    if not pages or is_probably_scanned(pages):
        try:
            pages_fb = extract_pages_text_pypdf(pdf_path)
            if sum(len((t or "").strip()) for t in pages_fb) > sum(len((t or "").strip()) for t in pages):
                pages = pages_fb
        except Exception as e:
            LOGGER.warning("pypdf fallback failed: %s", e)
    if enable_ocr and (not pages or is_probably_scanned(pages)):
        pages = ocr_pages(pdf_path, lang=ocr_lang, poppler_path=poppler_path,
                          tesseract_cmd=tesseract_cmd, tessdata_dir=tessdata_dir)
    return pages


# ----------------------------
# Credit parsing
# ----------------------------

@dataclass
class Contract:
    financing_type: Optional[str]
    creditor: Optional[str]
    contract_amount: Optional[int]
    periodic_payment: Optional[int]
    outstanding: Optional[int]
    current_overdue_days: Optional[int]
    current_overdue_amount: Optional[int]
    end_date: Optional[str]
    max_overdue_days: Optional[int]
    max_overdue_amount: Optional[int]
    cessionary: Optional[str]
    status: Optional[str]


SECTION_ACTIVE = "ДЕЙСТВУЮЩИЕ ДОГОВОРА"
SECTION_COMPLETED = "ЗАВЕРШЕННЫЕ ДОГОВОРЫ"
SECTION_OLD = "ДОГОВОРЫ, ЗАВЕРШЕННЫЕ БОЛЕЕ 5 ЛЕТ НАЗАД"
SECTION_REVOKED = "ОТОЗВАННЫЕ ДОГОВОРЫ"


def split_contract_blocks(text: str) -> List[str]:
    markers = list(re.finditer(r"\bКОНТРАКТ\s+\d+\b", text or ""))
    blocks: List[str] = []
    for i, m in enumerate(markers):
        start = m.start()
        end = markers[i + 1].start() if i + 1 < len(markers) else len(text)
        blocks.append(text[start:end])
    return blocks


def get_section(text: str, start: str, end: Optional[str]) -> str:
    full_text = text or ""
    starts = [m.start() for m in re.finditer(re.escape(start), full_text)]
    if not starts:
        return ""
    best_section = ""
    for sidx in starts:
        if end:
            eidx = full_text.find(end, sidx + len(start))
            section = full_text[sidx:] if eidx == -1 else full_text[sidx:eidx]
        else:
            section = full_text[sidx:]
        if re.search(r"\bКОНТРАКТ\s+\d+\b", section):
            best_section = section
    if best_section:
        return best_section
    sidx = starts[-1]
    if end:
        eidx = full_text.find(end, sidx + len(start))
        return full_text[sidx:] if eidx == -1 else full_text[sidx:eidx]
    return full_text[sidx:]


def normalize_fin_type(ft: Optional[str]) -> Optional[str]:
    if not ft:
        return None
    for t in ["Кредитная карта", "Займ", "Ипотека", "Кредит"]:
        if t.lower() in ft.lower():
            return t
    ft = ft.strip()
    return ft.split()[0] if ft else None


def clean_org_name(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = normalize_ws(s).strip(" :;,-")
    if not s or s in {"-", "—"}:
        return None
    s = re.split(
        r"\bДата\b|\bНаименование цессионария\b|\bКод контракта\b|\bСтатус договора\b|\bГодовая эффективная ставка\b|\bдоговора\s*:",
        s, maxsplit=1, flags=re.IGNORECASE,
    )[0].strip()
    s = re.split(
        r"\bБИН\b|\bКоличество транзакций\b|\bСтатус реабилитации\b|\bДЕТАЛИЗАЦИЯ\b",
        s, maxsplit=1, flags=re.IGNORECASE,
    )[0].strip()
    s = re.split(r"\b(?:Общая\s+сумма\s+кредита|Сумма|Информация\s+по\s+состоянию|Количество\s+дней\s+просрочки|Минимальный\s+платеж|Периодичность|Вознаграждение|Банк\s+второго\s+уровня)\b", s, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    s = re.split(r"\s+\d{2}\.\d{2}\.\d{4}\b|\s+\d{8,}\b", s, maxsplit=1)[0].strip()
    if not s or s in {"-", "—"}:
        return None
    return s


def sanitize_org_candidate(s: Optional[str]) -> Optional[str]:
    s = clean_org_name(s)
    if not s:
        return None
    s = re.sub(r'^\(?Кредитор\)?\s*:?\s*', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'^Источник\s+информации\s*(?:\(\s*Кредитор\s*\))?\s*:?\s*', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'^Наименование\s+цессионария\s*:?\s*', '', s, flags=re.IGNORECASE).strip()
    s = s.strip(" :;,-")
    return s or None


def is_valid_org_candidate(value: Optional[str]) -> bool:
    candidate = sanitize_org_candidate(value)
    if not candidate or candidate in {'-', '—'}:
        return False
    if re.fullmatch(r'[0-9 .-]+', candidate):
        return False
    if re.search(r'\d{8,}', candidate):
        return False
    if re.search(r'кредитор|источник\s+информации|информации\s+по|договор|периодич|состояние', candidate, re.IGNORECASE):
        return False
    if re.search(r'основной\s+долг|вознаграждение|платеж|просроч|непогаш|банк\s+второго(?:\s+уровня)?', candidate, re.IGNORECASE):
        return False
    if re.fullmatch(r'(?:кредит|финанс|банк|акционерное\s+общество|микрофинансовая(?:\s+организация)?)?[)"»]*', candidate, re.IGNORECASE):
        return False
    return looks_like_org_name(candidate)


def looks_like_org_name(value: Optional[str]) -> bool:
    if not value:
        return False
    s = clean_org_name(value)
    if not s:
        return False
    if re.fullmatch(r"[-—0-9. ]+", s):
        return False
    if re.search(r'сумма|платеж|просроч|непогаш|состоянию|вознаграждение|банк\s+второго(?:\s+уровня)?', s, re.IGNORECASE):
        return False
    return bool(re.search(
        r'\b(АО|ТОО|ИП|ФОНД|БАНК|BANK|МИКРОФИН|МФО|КОЛЛЕКТОР|ФИНАНС|КРЕДИТ|COLLECTION|CREDIT)\b|["«»]',
        s, flags=re.IGNORECASE,
    ))


def parse_creditor(block: str) -> Optional[str]:
    lines = [ln.strip() for ln in (block or "").splitlines() if ln is not None]
    for i, ln in enumerate(lines):
        if not re.search(r"Источник\s+информации", ln, flags=re.IGNORECASE):
            continue
        name_parts = []
        part = re.sub(r"^Источник\s+информации\s*", "", ln, flags=re.IGNORECASE).strip()
        if part:
            name_parts.append(part)
        for j in range(i + 1, min(i + 5, len(lines))):
            nxt = lines[j]
            stripped_creditor = re.sub(r'^\(?Кредитор\)?\s*:?\s*', '', nxt, flags=re.IGNORECASE).strip()
            if re.search(
                r"Дата|БИН|Код|Статус|ДОГОВОР|Контракт|Сумма|Информация\s+по\s+состоянию|Количество\s+дней\s+просрочки|Минимальный\s+платеж|Периодичность|Вознаграждение",
                nxt, re.IGNORECASE,
            ):
                break
            if stripped_creditor and stripped_creditor not in {"-", "—"} and stripped_creditor != nxt:
                name_parts.append(stripped_creditor)
                continue
            if nxt and nxt not in {"-", "—"}:
                name_parts.append(nxt)
            else:
                break
        name = " ".join(name_parts)
        name = sanitize_org_candidate(name)
        if name and len(name) > 3:
            return name
    m = re.search(
        r"Источник\s+информации\s*\(\s*Кредитор\s*\)\s*:?\s*(?P<name>[^\r\n]+)",
        block or "", flags=re.IGNORECASE,
    )
    return sanitize_org_candidate(m.group("name")) if m else None


def parse_contract_status(block: str) -> Optional[str]:
    lines = [normalize_ws(ln) for ln in (block or "").splitlines()]
    service_pat = re.compile(
        r"^(Наименование\s+цессионария|БИН|Дата|Код|Периодичность|Форма|ДЕТАЛИЗАЦИЯ|ДОПОЛНИТЕЛЬНАЯ\s+ИНФОРМАЦИЯ|Классификация\s+договора|Номинальная\s+ставка|Годовая\s+эффективная\s+ставка|Контракт|КОНТРАКТ)\b",
        flags=re.IGNORECASE,
    )
    for i, ln in enumerate(lines):
        m = re.match(r"^Статус\s+договора\s*:?\s*(.*)$", ln, flags=re.IGNORECASE)
        if not m:
            continue
        parts = []
        first = normalize_ws(m.group(1))
        if first:
            parts.append(first)
        for nxt in lines[i + 1:]:
            nxt = normalize_ws(nxt)
            if not nxt:
                if parts:
                    break
                continue
            if service_pat.search(nxt):
                break
            parts.append(nxt)
        status = normalize_ws(" ".join(parts)).strip(" :;,-")
        return status or None
    status = extract_between_or_none(
        block,
        r"Статус\s+договора\s*:?\s*",
        r"\bНаименование\s+цессионария\b|\bБИН\b|\bДата\b|\bКод\b|\bПериодичность\b|\bФорма\b|\bДЕТАЛИЗАЦИЯ\b|\bДОПОЛНИТЕЛЬНАЯ\s+ИНФОРМАЦИЯ\b|\bКлассификация\s+договора\b|\bНоминальная\s+ставка\b|\bГодовая\s+эффективная\s+ставка\b",
    )
    return normalize_ws(status).strip(" :;,-") if status else None


def is_cession_status(status: Optional[str]) -> bool:
    normalized = normalize_ws(status or "").lower()
    if not normalized:
        return False
    return (
        "уступка/переуступка права требования" in normalized
        or "переуступка" in normalized
        or "уступка" in normalized
    )


def parse_cessionary(block: str) -> Optional[str]:
    lines = [normalize_ws(ln) for ln in (block or "").splitlines()]
    stop_pat = re.compile(
        r"^(Источник\s+информации|ДЕТАЛИЗАЦИЯ|ДОПОЛНИТЕЛЬНАЯ\s+ИНФОРМАЦИЯ|Максимальное|Платежная\s+дисциплина|КОНТРАКТ|Контракт|СВОДКА)\b",
        flags=re.IGNORECASE,
    )
    service_value_pat = re.compile(
        r"^(?:Наименование\s+цессионария|БИН\s+цессионария|БИН|Дата(?:\s+последнего\s+платежа)?|Код|Статус|Периодичность|Форма|Общее\s+количество\s+платежей|Кол-?во\s+непогашенных.*|Количество\s+непогашенных.*)\s*:?\s*(.*)$",
        flags=re.IGNORECASE,
    )
    service_only_pat = re.compile(
        r"^(?:БИН\s+цессионария|БИН|Дата(?:\s+последнего\s+платежа)?|Код|Статус|Периодичность|Форма|Общее\s+количество\s+платежей|Кол-?во\s+непогашенных.*|Количество\s+непогашенных.*)\s*:?(?:\s*[-—0-9.]*)$",
        flags=re.IGNORECASE,
    )

    def collect_from_lines(start_idx: int, seed: Optional[str]) -> Optional[str]:
        parts = []
        seed = sanitize_org_candidate(seed)
        if is_valid_org_candidate(seed):
            parts.append(seed)
        for nxt in lines[start_idx + 1:]:
            nxt = normalize_ws(nxt)
            if not nxt:
                if parts:
                    break
                continue
            if stop_pat.search(nxt):
                break
            service_match = service_value_pat.match(nxt)
            candidate = sanitize_org_candidate(service_match.group(1) if service_match else nxt)
            if service_only_pat.match(nxt):
                continue
            if not is_valid_org_candidate(candidate):
                if parts:
                    break
                continue
            parts.append(candidate)
        name = sanitize_org_candidate(" ".join(parts))
        if not is_valid_org_candidate(name):
            return None
        if re.search(r'вознаграждение|банк\s+второго', name or "", re.IGNORECASE):
            return None
        return name

    for i, ln in enumerate(lines):
        m = re.match(r"^Наименование\s+цессионария\s*:?\s*(.*)$", ln, flags=re.IGNORECASE)
        if not m:
            continue
        name = collect_from_lines(i, m.group(1))
        if name:
            return name
        for candidate_line in lines[max(0, i - 4):min(len(lines), i + 8)]:
            if candidate_line == ln:
                continue
            if re.search(r"Источник\s+информации|\(Кредитор\)|Статус\s+договора|Классификация\s+договора", candidate_line, re.IGNORECASE):
                continue
            candidate = sanitize_org_candidate(candidate_line)
            if is_valid_org_candidate(candidate):
                return candidate
    return None


def has_real_cession(c: Contract) -> bool:
    if not c.cessionary:
        return False
    if not c.creditor:
        return True
    return normalize_ws(c.cessionary).lower() != normalize_ws(c.creditor).lower()


def contract_has_cession(c: Contract) -> bool:
    return has_real_cession(c) or is_cession_status(c.status)


def is_valid_completed(c: Contract) -> bool:
    return (
        (c.max_overdue_days or 0) > 0 or
        (c.max_overdue_amount or 0) > 0 or
        contract_has_cession(c)
    )


def extract_between_or_none(text: str, start_pat: str, end_pat: str) -> Optional[str]:
    m = re.search(start_pat, text or "", flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    start = m.end()
    m2 = re.search(end_pat, (text or "")[start:], flags=re.IGNORECASE | re.DOTALL)
    val = (text or "")[start:] if not m2 else (text or "")[start:start + m2.start()]
    val = normalize_ws(val)
    if not val or val in {"-", "—"}:
        return None
    return val


def parse_current_overdue_days(block: str) -> Optional[int]:
    m = re.search(
        r"Количество\s+дней\s+просрочки\s*:\s*([\d\s]+)\s*Пеня",
        block or "", flags=re.IGNORECASE,
    )
    if m:
        return parse_int_with_spaces(m.group(1))
    m = re.search(r"Количество\s+дней\s+просрочки\s*:\s*([\d\s]+)", block or "", flags=re.IGNORECASE)
    return parse_int_with_spaces(m.group(1)) if m else None


def parse_current_overdue_amount(block: str) -> Optional[int]:
    m = re.search(
        r"Сумма\s+просроченных\s+взносов\s*:\s*([\d\s.,]+)\s*KZT\s*Штраф",
        block or "", flags=re.IGNORECASE,
    )
    if m:
        return money_to_int_kzt(m.group(1))
    m = re.search(
        r"Сумма\s+просроченных\s+взносов\s*:\s*([\d\s.,]+)\s*KZT",
        block or "", flags=re.IGNORECASE,
    )
    return money_to_int_kzt(m.group(1)) if m else None


def parse_max_overdue_days(block: str) -> Optional[int]:
    nums: List[int] = []
    nums += [
        parse_int_with_spaces(x)
        for x in re.findall(
            r"Максимальное\s+количество\s+дней\s+просрочки[^:]*:\s*([\d\s]+)",
            block or "", flags=re.IGNORECASE,
        )
    ]
    nums = [n for n in nums if n is not None]
    return max(nums) if nums else None


def parse_max_overdue_amount(block: str) -> Optional[int]:
    nums: List[int] = []
    nums += [
        money_to_int_kzt(x)
        for x in re.findall(
            r"Максимальная\s+сумма\s+просроченных\s+взносов[^:]*:\s*([\d\s.,]+)\s*KZT",
            block or "", flags=re.IGNORECASE,
        )
    ]
    nums = [n for n in nums if n is not None]
    return max(nums) if nums else None


def parse_contract_block(block: str) -> Contract:
    creditor = parse_creditor(block)

    fin_type_raw = extract_between_or_none(
        block,
        r"Вид финансирования:\s*",
        r"\bИсточник финансирования\b|\bРоль субъекта\b|\bЦель кредита\b|\bДОГОВОР\b",
    )
    fin_type = normalize_fin_type((fin_type_raw or "").splitlines()[0].strip() if fin_type_raw else None)

    limit = extract_first(block, r"Сумма кредитного лимита\s*:?\s*([\d\s.,]+)\s?KZT")
    total_credit = extract_first(block, r"Общая сумма кредита\s*/\s*валюта\s*:\s*([\d\s.,]+)\s*KZT")
    contract_amount = money_to_int_kzt(limit) if limit else (money_to_int_kzt(total_credit) if total_credit else None)

    min_payment = extract_first(block, r"Минимальный платеж\s*:\s*([\d\s.,]+)\s*KZT")
    periodic = extract_first(block, r"Сумма периодического платежа\s*:\s*([\d\s.,]+)\s*KZT")
    periodic_payment = money_to_int_kzt(min_payment) if min_payment else (money_to_int_kzt(periodic) if periodic else None)

    used = extract_first(block, r"Использованная сумма\s*\(подлежащая погашению\)\s*:\s*([\d\s.,]+)\s*KZT")
    unpaid = extract_first(block, r"Непогашенная сумма по кредиту\s*:\s*([\d\s.,]+)\s*KZT")
    outstanding = money_to_int_kzt(used) if used else (money_to_int_kzt(unpaid) if unpaid else None)

    current_overdue_days = parse_current_overdue_days(block)
    current_overdue_amount = parse_current_overdue_amount(block)

    end_date = extract_first(block, r"Дата фактического завершения\s*:\s*(\d{2}\.\d{2}\.\d{4}|-)")
    end_date = end_date if (end_date and end_date != "-") else None

    max_overdue_days = parse_max_overdue_days(block)
    max_overdue_amount = parse_max_overdue_amount(block)

    status = parse_contract_status(block)
    cessionary = parse_cessionary(block)

    if is_cession_status(status):
        LOGGER.debug(
            "Detected cession contract: creditor=%r cessionary=%r status=%r",
            creditor, cessionary, status,
        )

    # For active contracts: if current overdue is 0 but max > 0, use max
    if end_date is None:
        if (current_overdue_days or 0) == 0 and (max_overdue_days or 0) > 0:
            current_overdue_days = max_overdue_days
        if (current_overdue_amount or 0) == 0 and (max_overdue_amount or 0) > 0:
            current_overdue_amount = max_overdue_amount

    return Contract(
        financing_type=fin_type,
        creditor=creditor,
        contract_amount=contract_amount,
        periodic_payment=periodic_payment,
        outstanding=outstanding,
        current_overdue_days=current_overdue_days,
        current_overdue_amount=current_overdue_amount,
        end_date=end_date,
        max_overdue_days=max_overdue_days,
        max_overdue_amount=max_overdue_amount,
        cessionary=cessionary,
        status=status,
    )


def parse_credit_header(full_text: str) -> Dict[str, Any]:
    pkr_s = extract_first(full_text, r"\bПКР\s*[:：]\s*(\d{1,4})\b")
    pkr = int(pkr_s) if pkr_s else None

    generated_at = None
    m = re.search(
        r"\bПОЛН[А-Я\s]*КРЕДИТНЫЙ ОТЧЕТ\b.*?\b(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2}:\d{2})\b",
        full_text or "", flags=re.IGNORECASE | re.DOTALL,
    )
    if m:
        generated_at = f"{m.group(1)} {m.group(2)}"

    iin = extract_first(full_text, r"\bИИН\s*/\s*БИН\s*[:：]\s*(\d{12})\b")
    if not iin:
        iin = extract_first(full_text, r"\bИИН\s*[:：]\s*(\d{12})\b")

    labels = {
        "active_no_overdue": "Действующие договоры без просрочки",
        "active_overdue": "Действующие договоры с просрочкой",
        "ended_no_overdue": "Завершенные договоры без просрочки",
        "ended_overdue": "Завершенные договоры с просрочкой",
        "ended_5y": "Завершенные более 5 лет назад",
        "revoked": "Отозванные договоры",
    }
    counts: Dict[str, Optional[int]] = {}
    for k, label in labels.items():
        m2 = re.search(rf"(\d+)\s+{re.escape(label)}", full_text or "")
        counts[k] = int(m2.group(1)) if m2 else None

    return {"pkr": pkr, "generated_at": generated_at, "iin": iin, "counts": counts}


def extract_credit_report(
    pdf_path: Path,
    enable_ocr: bool,
    ocr_lang: str,
    poppler_path: Optional[str],
    tesseract_cmd: Optional[str],
    tessdata_dir: Optional[str],
    plumber_layout: bool = False,
) -> Dict[str, Any]:
    pages = extract_pages_text(
        pdf_path,
        enable_ocr=enable_ocr,
        ocr_lang=ocr_lang,
        poppler_path=poppler_path,
        tesseract_cmd=tesseract_cmd,
        tessdata_dir=tessdata_dir,
        plumber_layout=plumber_layout,
    )
    if not pages:
        raise RuntimeError("Не удалось извлечь текст из кредитного PDF")

    full = "\n\n".join(pages)
    header = parse_credit_header(full)

    active_sec = get_section(full, SECTION_ACTIVE, SECTION_COMPLETED)
    completed_sec = get_section(full, SECTION_COMPLETED, SECTION_OLD)
    old_sec = get_section(full, SECTION_OLD, SECTION_REVOKED)
    revoked_sec = get_section(full, SECTION_REVOKED, None)

    active = [parse_contract_block(b) for b in split_contract_blocks(active_sec)]
    active = [c for c in active if c.end_date is None]

    completed = [parse_contract_block(b) for b in split_contract_blocks(completed_sec) if b.strip()]
    completed = [c for c in completed if c.end_date]

    old = [parse_contract_block(b) for b in split_contract_blocks(old_sec) if b.strip()]
    old = [c for c in old if c.end_date]

    completed = [c for c in completed if is_valid_completed(c)]
    old = [c for c in old if is_valid_completed(c)]
    revoked = [parse_contract_block(b) for b in split_contract_blocks(revoked_sec) if b.strip()]
    revoked = [c for c in revoked if c.end_date]

    return {
        "pkr": header.get("pkr"),
        "generated_at": header.get("generated_at"),
        "iin": header.get("iin"),
        "counts": header.get("counts", {}),
        "active": active,
        "completed_recent": completed,
        "completed_old": old,
        "revoked": revoked,
    }


def contract_has_any_overdue(c: Contract) -> bool:
    return any([
        (c.current_overdue_days or 0) > 0,
        (c.current_overdue_amount or 0) > 0,
        (c.max_overdue_days or 0) > 0,
        (c.max_overdue_amount or 0) > 0,
    ])


def mask_iin(iin: Optional[str]) -> str:
    if not iin:
        return ""
    if len(iin) != 12:
        return iin
    return "********" + iin[-4:]


# ============================================================
# ГКБ / МКБ ("Мемлекеттік кредиттік бюро" / Государственное
# кредитное бюро) — второй формат кредитного отчёта.
# Текст на казахском, слова в PDF слипаются без пробелов,
# каждый договор ("Міндеттеме N") на отдельной странице.
# На выходе — та же структура, что и у extract_credit_report.
# ============================================================

# Метки правой колонки: обрезают "прилипший" из-за 2-колоночной вёрстки текст.
_GKB_RIGHT_COL = re.compile(
    r"Шарттың|БИН[:]|Келісімшарт|Кредиттің|Кредитбер|Қаржыландыру|Айсайынғы|"
    r"Мерзімі|Алдағы|Нақты|Өтелмеген|Жарналарды|Номиналды|Жылдық|Төлемдерді|"
    r"Ата-ана|Субъектіні|ЖСН"
)


def _gkb_money(s: Optional[str]) -> Optional[int]:
    if not s:
        return None
    s = s.replace(" ", "").replace(" ", "").replace(",", "")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    return int(float(m.group(1))) if m else None


def _gkb_is_nodata(v: Optional[str]) -> bool:
    return (not v) or ("Деректержоқ" in v) or ("Деректер жоқ" in v) or v in {"-", "—"}


def _gkb_val_after(label: str, block: str) -> Optional[str]:
    m = re.search(re.escape(label) + r"\s*([^\r\n]*)", block or "")
    return m.group(1).strip() if m else None


def _gkb_date_after(label: str, block: str) -> Optional[str]:
    m = re.search(re.escape(label) + r"\s*(\d{2}\.\d{2}\.\d{4})", block or "")
    return m.group(1) if m else None


def _gkb_cut_right(s: Optional[str]) -> Optional[str]:
    if not s:
        return s
    return _GKB_RIGHT_COL.split(s)[0].strip()


def _gkb_balanced(s: str) -> bool:
    return s.count("(") <= s.count(")") and s.count('"') % 2 == 0


# Известные банки по БИН — точные названия (в отчёте ГКБ текст слипается без пробелов)
_GKB_BANKS = {
    "971240001315": "АО «Kaspi Bank»",
    "920140000084": "АО «Alatau City Bank»",
    "951140000151": "АО «АТФБанк» (ныне Jusan Bank)",
    "980640000093": "АО «Банк ЦентрКредит»",
    "930540000147": "АО «Home Credit Bank»",
    "940140000385": "АО «Народный Банк Казахстана»",
    "990740000683": "АО «ForteBank»",
    "060140002873": "АО «Bereke Bank»",
    "070140000075": "АО «Jusan Bank»",
    "941240000108": "АО «Банк ЦентрКредит»",
}


def _gkb_clean_bank(name: Optional[str], bin_: Optional[str] = None) -> Optional[str]:
    """Чистит склеенное название банка из ГКБ. По БИН — точное имя, иначе — расстановка пробелов."""
    if bin_ and bin_ in _GKB_BANKS:
        return _GKB_BANKS[bin_]
    if not name:
        return name
    s = name
    s = re.sub(r'(АО|ТОО|АҚ|ЖШС)\s*"', r"\1 «", s)          # префикс + открывающая кавычка
    s = s.replace('"', "»")                                  # прочие кавычки → закрывающая
    s = re.sub(r"(?<=\S)\(", " (", s)                        # пробел перед (
    s = s.replace("Дочернийбанк", "Дочерний банк").replace("ДБАО", "ДБ АО")
    s = re.sub(r"(?<=[a-zа-яё])(?=[A-ZА-ЯЁ])", " ", s)       # разделить слипшийся CamelCase
    s = re.sub(r"»(?=[A-ZА-ЯЁ])", "» «", s)                  # «...»(След → «...» «След
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_gkb_blocks(section: str) -> List[str]:
    """Режет секцию на блоки по маркерам 'Міндеттеме1', 'Міндеттеме2'..."""
    marks = list(re.finditer(r"(?m)^Міндеттеме\d+", section or ""))
    out: List[str] = []
    for i, m in enumerate(marks):
        s = m.start()
        e = marks[i + 1].start() if i + 1 < len(marks) else len(section)
        out.append(section[s:e])
    return out


def parse_gkb_block(block: str, phase: str) -> Contract:
    """Разбирает один договор ГКБ. phase: 'active' | 'completed'."""
    lines = block.splitlines()

    creditor = _gkb_cut_right(_gkb_val_after("Кредитор:", block))
    # длинное название переносится на следующую строку ("...Jysan" / "Bank")")
    if creditor and not _gkb_balanced(creditor):
        for i, ln in enumerate(lines):
            if ln.strip().startswith("Кредитор:") and i + 1 < len(lines):
                cont = lines[i + 1].strip().split(" ")[0]
                if cont and not cont.startswith(("Шарт", "БИН", "Кредит", "Келі")):
                    creditor = (creditor + " " + cont).strip()
                break
    if _gkb_is_nodata(creditor):
        creditor = None

    # БИН для точного названия банка по справочнику
    bin_ = None
    mb = re.search(r"БИН:\s*(\d{6,})", block)
    if mb:
        bin_ = mb.group(1)
    creditor = _gkb_clean_bank(creditor, bin_)

    fin_raw = (_gkb_val_after("Қаржыландырутүрі:", block) or "").lower()
    if "карт" in fin_raw:
        fin = "Кредитная карта"
    elif "ипотек" in fin_raw:
        fin = "Ипотека"
    elif "қарыз" in fin_raw:
        fin = "Займ"
    else:
        fin = fin_raw or None

    total = _gkb_money(_gkb_val_after("Шарттыңжалпысомасы/валюта:", block))
    monthly = _gkb_money(_gkb_val_after("Айсайынғытөлемсомасы/валюта:", block))
    upcoming = _gkb_money(_gkb_val_after("Алдағытөлемдерсомасы/валюта", block))
    cur_over_amt = _gkb_money(_gkb_val_after("Мерзіміөткенжарналарсомасы/валюта:", block))

    cur_over_days = None
    m = re.search(r"Мерзіміөткенкүндерсаны:\s*(\d+)", block)
    if m:
        cur_over_days = int(m.group(1))

    start_date = _gkb_date_after("Келісімшарттыңқолданылумерзімініңбасталукүні:", block)
    end_contract = _gkb_date_after("Келісімшарттыңқолданылумерзімініңаяқталукүні:", block)
    actual_paid = _gkb_date_after("Нақтыөтелгенкүні:", block)

    # период договора (метка может быть склеена со значением)
    m = re.search(r"Келісімшарткезеңі\s*([^\r\n]*)", block)
    period_raw = (m.group(1).strip() if m else "").lower()

    # цессия: статус "...талап ету құқықтарын басқа тұлғаға табыстау"
    # или период "Кредиторды ауыстыру"
    bl = block.lower()
    is_cession = (
        "табыста" in bl or "ауыстыру" in period_raw or "талапетуқұқық" in bl
    )
    status = "Уступка/переуступка права требования" if is_cession else "Стандартный"

    # цессионарий: значение на той же или следующей строке
    cessionary = None
    m = re.search(
        r"құқық \(талаптар\) берілген тұлғаның атауы:\s*([^\r\n]*)(?:\r?\n\s*([^\r\n]*))?",
        block,
    )
    if m:
        for cand in (m.group(1), m.group(2)):
            cand = (cand or "").strip()
            if (cand and not _gkb_is_nodata(cand) and not cand.startswith("Шарт")
                    and re.search(r'[«"]|АО|ТОО|Bank|БАНК|Jusan|Jysan|Финанс', cand)):
                cessionary = _gkb_clean_bank(cand.strip())
                break

    # макс. просрочка — из помесячной сетки на странице истории платежей
    max_days = 0
    max_amt = 0
    for line in lines:
        line = line.strip()
        md = re.match(r"^өткен\s+(.+)$", line)
        if md and "сома" not in line:
            for tok in md.group(1).split():
                t = tok.replace(",", "")
                if t.isdigit():
                    max_days = max(max_days, int(t))
        ma = re.match(r"^өткенсома\s+(.+)$", line)
        if ma:
            for tok in ma.group(1).split():
                v = _gkb_money(tok)
                if v:
                    max_amt = max(max_amt, v)

    end_date = None
    if phase == "completed":
        end_date = actual_paid or end_contract

    return Contract(
        financing_type=fin,
        creditor=creditor,
        contract_amount=total,
        periodic_payment=monthly,
        outstanding=upcoming,
        current_overdue_days=cur_over_days,
        current_overdue_amount=cur_over_amt,
        end_date=end_date,
        max_overdue_days=max_days or None,
        max_overdue_amount=max_amt or None,
        cessionary=cessionary,
        status=status,
    )


def _gkb_fio(full: str) -> Optional[str]:
    teg = _gkb_val_after("Тегі:", full)
    at = _gkb_val_after("Аты:", full)
    m = re.search(r"Әкесінің\s*аты\s*:?\s*(\S+)", full)
    ak = m.group(1) if m else None
    parts = [p.split()[0] for p in (teg, at, ak) if p and not _gkb_is_nodata(p)]
    fio = " ".join(parts).strip()
    if fio:
        return fio
    # запасной вариант — ФИО из реквизитов (латиница отправителя/получателя)
    m = re.search(r"ИИН/БИН\s+\d{12}\s*\n\s*([А-ЯЁЀ-ӿ]+(?:\s+[А-ЯЁЀ-ӿ]+){1,2})", full)
    return m.group(1).strip() if m else None


def extract_gkb_report(
    pdf_path: Path,
    enable_ocr: bool = False,
    ocr_lang: str = "kaz+rus",
    poppler_path: Optional[str] = None,
    tesseract_cmd: Optional[str] = None,
    tessdata_dir: Optional[str] = None,
    plumber_layout: bool = False,
) -> Dict[str, Any]:
    pages = extract_pages_text(
        pdf_path,
        enable_ocr=enable_ocr,
        ocr_lang=ocr_lang,
        poppler_path=poppler_path,
        tesseract_cmd=tesseract_cmd,
        tessdata_dir=tessdata_dir,
        plumber_layout=plumber_layout,
    )
    if not pages:
        raise RuntimeError("Не удалось извлечь текст из кредитного PDF (ГКБ)")
    full = "\n".join(pages)

    iin = None
    m = re.search(r"ЖСН:\s*(\d{12})", full) or re.search(r"ИИН/БИН\s+(\d{12})", full)
    if m:
        iin = m.group(1)

    fio = _gkb_fio(full)

    generated_at = None
    m = re.search(r"Берілгенкүні:\s*(\d{2}\.\d{2}\.\d{4})", full)
    if m:
        gd = m.group(1)
        mt = re.search(r"Берілгенуақыты:\s*(\d{2}:\d{2})", full)
        generated_at = f"{gd} {mt.group(1)}" if mt else gd

    # секции: активные / завершённые
    ai = full.find("ҚОЛДАНЫСТАҒЫШАРТТАРБОЙЫНША")
    ci = full.find("АЯҚТАЛҒАНШАРТТАРТУРАЛЫ")
    ei = full.find("КРЕДИТТІКТАРИХСУБЪЕКТІСІТУРАЛЫАҒЫМДАҒЫ")
    if ei == -1:
        ei = full.find("МАҢЫЗДЫАҚПАРАТ")
    active_sec = full[ai:ci] if (ai != -1 and ci != -1) else (full[ai:] if ai != -1 else "")
    completed_sec = full[ci:ei] if (ci != -1 and ei != -1) else (full[ci:] if ci != -1 else "")

    active = [parse_gkb_block(b, "active") for b in split_gkb_blocks(active_sec)]
    active = [c for c in active if c.end_date is None]

    completed_all = [parse_gkb_block(b, "completed") for b in split_gkb_blocks(completed_sec)]
    completed_all = [c for c in completed_all if is_valid_completed(c)]

    # разбивка завершённых на <=5 лет и >5 лет по дате завершения
    ref = None
    if generated_at:
        dm = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", generated_at)
        if dm:
            from datetime import date
            ref = date(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
    if ref is None:
        from datetime import date
        ref = date.today()

    def years_ago(c: Contract) -> float:
        if not c.end_date:
            return 0.0
        dm = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", c.end_date)
        if not dm:
            return 0.0
        from datetime import date
        end = date(int(dm.group(3)), int(dm.group(2)), int(dm.group(1)))
        return (ref - end).days / 365.25

    recent = [c for c in completed_all if years_ago(c) <= 5]
    old = [c for c in completed_all if years_ago(c) > 5]

    n_active = re.search(r"Қолданыстағыміндеттемелер\((\d+)\)", full)
    n_completed = re.search(r"Аяқталғанміндеттемелер\((\d+)\)", full)
    counts = {
        "active_no_overdue": None,
        "active_overdue": None,
        "ended_no_overdue": None,
        "ended_overdue": None,
        "ended_5y": None,
        "revoked": None,
        "active_total": int(n_active.group(1)) if n_active else None,
        "completed_total": int(n_completed.group(1)) if n_completed else None,
    }

    return {
        "source": "gkb",
        "pkr": None,  # ГКБ не выводит ПКР
        "generated_at": generated_at,
        "iin": iin,
        "fio": fio,
        "counts": counts,
        "active": active,
        "completed_recent": recent,
        "completed_old": old,
        "revoked": [],
    }


def detect_report_type(pdf_path: Path,
                       enable_ocr: bool = False,
                       ocr_lang: str = "rus",
                       poppler_path: Optional[str] = None,
                       tesseract_cmd: Optional[str] = None,
                       tessdata_dir: Optional[str] = None) -> str:
    """Определяет бюро по тексту первых страниц: 'pkb' | 'gkb' | 'unknown'."""
    try:
        pages = extract_pages_text(
            pdf_path, enable_ocr=enable_ocr, ocr_lang=ocr_lang,
            poppler_path=poppler_path, tesseract_cmd=tesseract_cmd,
            tessdata_dir=tessdata_dir, plumber_layout=False,
        )
    except Exception:
        return "unknown"
    head = "\n".join(pages[:3])
    low = head.lower()
    gkb_markers = ["жеке кредиттік есеп", "мемлекеттіккредиттікбюро",
                   "государственное кредитное бюро", "mkb.kz", "міндеттеме"]
    pkb_markers = ["персональный кредитный отчет", "1cb.kz",
                   "первое кредитное бюро"]
    if any(x in low for x in gkb_markers):
        return "gkb"
    if any(x in low for x in pkb_markers) or re.search(r"\bПКР\s*[:：]", head):
        return "pkb"
    # запасная эвристика по всему тексту
    full = "\n".join(pages).lower()
    if "міндеттеме" in full or "mkb.kz" in full:
        return "gkb"
    if "пкр" in full or "1cb.kz" in full:
        return "pkb"
    return "unknown"


def extract_any_credit_report(
    pdf_path: Path,
    enable_ocr: bool,
    ocr_lang: str,
    poppler_path: Optional[str],
    tesseract_cmd: Optional[str],
    tessdata_dir: Optional[str],
    plumber_layout: bool = False,
) -> Dict[str, Any]:
    """Единая точка входа: сама определяет ПКБ или ГКБ и парсит соответственно."""
    kind = detect_report_type(
        pdf_path, enable_ocr=enable_ocr, ocr_lang=ocr_lang,
        poppler_path=poppler_path, tesseract_cmd=tesseract_cmd,
        tessdata_dir=tessdata_dir,
    )
    if kind == "gkb":
        LOGGER.info("Тип отчёта: ГКБ/МКБ")
        return extract_gkb_report(
            pdf_path, enable_ocr=enable_ocr, ocr_lang="kaz+rus",
            poppler_path=poppler_path, tesseract_cmd=tesseract_cmd,
            tessdata_dir=tessdata_dir, plumber_layout=plumber_layout,
        )
    LOGGER.info("Тип отчёта: ПКБ (1cb.kz)%s", "" if kind == "pkb" else " [по умолчанию]")
    data = extract_credit_report(
        pdf_path, enable_ocr=enable_ocr, ocr_lang=ocr_lang,
        poppler_path=poppler_path, tesseract_cmd=tesseract_cmd,
        tessdata_dir=tessdata_dir, plumber_layout=plumber_layout,
    )
    data.setdefault("source", "pkb")
    data.setdefault("fio", None)
    return data


# ----------------------------
# Pension parsing + per-org СРЗП
# ----------------------------

@dataclass
class PensionRow:
    date: str
    amount: Optional[int]
    period: Optional[str]
    sender: Optional[str]
    sender_bin: Optional[str]
    status: Optional[str]


def parse_pension_pdf_tables(pdf_path: Path) -> Dict[str, Any]:
    import pdfplumber

    with pdfplumber.open(str(pdf_path)) as pdf:
        pages = list(pdf.pages)
        full_text = "\n\n".join((p.extract_text() or "") for p in pages)

        fio = extract_first(full_text, r"ТАӘ/ФИО\s*:\s*([^\n\r]+)")
        iin = extract_first(full_text, r"ЖСН/ИИН\s*:\s*(\d{12})")
        period = extract_first(full_text, r"Период\s*:\s*([0-9.\- ]+)")
        doc_no = extract_first(full_text, r"№\s*:\s*([0-9]{6,})")

        rows: List[PensionRow] = []
        for page in pages:
            for table in (page.extract_tables() or []):
                if not table or len(table) < 2:
                    continue
                header = table[0]
                header_join = " | ".join([(h or "") for h in header])
                if "Сумма" not in header_join:
                    continue
                for r in table[1:]:
                    if not r or len(r) < 11:
                        continue
                    date_in = (r[0] or "").strip()
                    if not re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", date_in):
                        continue
                    amount_s = (r[8] or "").strip()
                    try:
                        amount_i = int(float(amount_s.replace(" ", "").replace("\u00a0", "")))
                    except Exception:
                        amount_i = None
                    rows.append(PensionRow(
                        date=date_in,
                        amount=amount_i,
                        period=normalize_ws(r[10] or "") or None,
                        sender=normalize_ws(r[5] or "") or None,
                        sender_bin=normalize_ws(r[6] or "") or None,
                        status=normalize_ws(r[9] or "") or None,
                    ))

    org = None
    if rows and any(rr.sender for rr in rows):
        org = Counter([rr.sender for rr in rows if rr.sender]).most_common(1)[0][0]

    return {"fio": fio, "iin": iin, "period": period, "doc_no": doc_no, "rows": rows, "org": org}


def calc_pension_avg_vals(vals: List[float], coef: float = 7.9) -> int:
    """
    Формула пользователя:
      СРЗП = (sum*7.9/6 + (sum-min-max)*7.9/(n-2)) / 2
    Если n <= 2: вторая часть = sum*7.9/n
    """
    if not vals:
        return 0
    n = len(vals)
    total = sum(vals)
    mn = min(vals)
    mx = max(vals)

    part1 = total * coef / 6
    if n > 2:
        part2 = (total - mn - mx) * coef / (n - 2)
    else:
        part2 = total * coef / n

    return round((part1 + part2) / 2)


def group_pension_by_org(rows: List[PensionRow]) -> Dict[str, List[PensionRow]]:
    """Группировка по BIN (sender_bin) — один БИН = одна организация.
    Если БИН отсутствует — fallback на название."""
    groups: Dict[str, List[PensionRow]] = defaultdict(list)
    for rr in rows:
        key = rr.sender_bin or rr.sender or "Неизвестно"
        groups[key].append(rr)
    return dict(groups)


def get_org_display_name(bin_key: str, rows: List["PensionRow"]) -> str:
    """Возвращает читаемое название организации по её БИН-ключу."""
    for rr in rows:
        if (rr.sender_bin or rr.sender or "Неизвестно") == bin_key and rr.sender:
            return normalize_ws(rr.sender)
    return bin_key


# ----------------------------
# Validation / logging
# ----------------------------

def validate_credit_data(data: Dict[str, Any], strict: bool) -> None:
    pkr = data.get("pkr")
    if pkr is None:
        msg = "ПКР не найден (pkr=None). Проверьте извлечение текста/якоря."
        if strict:
            raise ValueError(msg)
        LOGGER.warning(msg)

    counts = data.get("counts") or {}
    if counts.get("active_no_overdue") is not None and counts.get("active_overdue") is not None:
        expected = int(counts["active_no_overdue"]) + int(counts["active_overdue"])
        actual = len(data.get("active") or [])
        if expected != actual:
            msg = f"Несовпадение активных договоров: header={expected}, parsed={actual}"
            if strict:
                raise ValueError(msg)
            LOGGER.warning(msg)

    active = data.get("active") or []
    if active and sum(1 for c in active if not c.creditor) / len(active) >= 0.5:
        LOGGER.warning("У >=50%% активных договоров не распознан кредитор.")

    if len(active) > 18:
        LOGGER.warning("Активных договоров %s, но шаблон вмещает 18: лишние будут отброшены.", len(active))
    if len(data.get("completed_recent") or []) > 10:
        LOGGER.warning("Завершённых (<=5 лет) %s, но шаблон вмещает 10.", len(data.get("completed_recent") or []))
    if len(data.get("completed_old") or []) > 10:
        LOGGER.warning("Завершённых (>5 лет) %s, но шаблон вмещает 10.", len(data.get("completed_old") or []))


def validate_pension_data(pension: Dict[str, Any], strict: bool) -> None:
    if not pension.get("fio") or not pension.get("iin"):
        msg = "Пенсионный документ: не найдено ФИО и/или ИИН."
        if strict:
            raise ValueError(msg)
        LOGGER.warning(msg)
    rows: List[PensionRow] = pension.get("rows") or []
    if not rows:
        msg = "Пенсионный документ: не извлечены строки таблицы."
        if strict:
            raise ValueError(msg)
        LOGGER.warning(msg)


def validate_cross_iin(credit_iin: Optional[str], pension_iin: Optional[str], strict: bool) -> None:
    if not credit_iin or not pension_iin:
        return
    if credit_iin != pension_iin:
        msg = (
            f"ИИН кредитного отчёта ({mask_iin(credit_iin)}) и пенсионной выписки ({mask_iin(pension_iin)}) НЕ совпадают."
        )
        if strict:
            raise ValueError(msg)
        LOGGER.warning(msg)


# ----------------------------
# Safe cell writer (handles merged cells)
# ----------------------------

def safe_set(ws, cell_ref: str, value) -> None:
    """Write to a cell, handling merged cells by writing only to top-left."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import coordinate_to_tuple

    cell = ws[cell_ref]

    # If it's a MergedCell instance, find the top-left of the merge range
    if isinstance(cell, MergedCell):
        for mc in ws.merged_cells.ranges:
            min_col = mc.min_col
            min_row = mc.min_row
            max_col = mc.max_col
            max_row = mc.max_row
            row, col = coordinate_to_tuple(cell_ref)
            if min_row <= row <= max_row and min_col <= col <= max_col:
                ws.cell(row=min_row, column=min_col).value = value
                return
        # Fallback: write to top-left via coordinates
        row, col = coordinate_to_tuple(cell_ref)
        ws.cell(row=row, column=col).value = value
    else:
        cell.value = value


# ----------------------------
# LEGACY WRITER (sheet "Кредиты")
# ----------------------------

def write_output_legacy_excel(
    template_path: Path,
    output_path: Path,
    pkr: Optional[int],
    active: List[Contract],
    recent: List[Contract],
    old: List[Contract],
    revoked: Optional[List[Contract]] = None,
    blank_zero: bool = True,
    only_loans_active: bool = False,
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(str(template_path))
    if "Кредиты" not in wb.sheetnames:
        raise KeyError("В шаблоне нет листа 'Кредиты'")
    ws = wb["Кредиты"]

    if only_loans_active:
        active = [c for c in active if c.financing_type == "Займ"]

    n_active    = len(active)
    n_completed = len(recent) + len(old)
    n_revoked   = len(revoked or [])
    n_cession   = sum(1 for c in (recent + old) if contract_has_cession(c))
    pkr_str     = str(pkr) if pkr is not None else "—"

    ws["A1"].value = (
        f"ПКО:  Действ {n_active}   "
        f"Завершённый {n_completed}   "
        f"Стандартный {n_active + n_completed}   "
        f"Уступка {n_cession}   "
        f"Отозванный {n_revoked}   "
        f"ПКР {pkr_str}"
    )

    # Active: rows 6..23 (18 rows) — headers occupy rows 4-5
    for row in range(6, 24):
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            safe_set(ws, f"{col}{row}", None)

    for i, c in enumerate(active[:18]):
        row = 6 + i
        safe_set(ws, f"B{row}", c.financing_type)
        safe_set(ws, f"C{row}", c.creditor)
        safe_set(ws, f"D{row}", to_cell_str(c.contract_amount, blank_zero=False))
        safe_set(ws, f"E{row}", to_cell_str(c.periodic_payment, blank_zero=blank_zero))
        safe_set(ws, f"F{row}", to_cell_str(c.outstanding, blank_zero=blank_zero))
        safe_set(ws, f"G{row}", to_cell_str(c.current_overdue_days, blank_zero=blank_zero))
        safe_set(ws, f"H{row}", to_cell_str(c.current_overdue_amount, blank_zero=blank_zero))

    # Completed (<= 5 years) + Revoked: rows 30..39 (10 rows)
    # Отозванные идут вместе с завершёнными, с пометкой "[Отозванный]"
    revoked_list = list(revoked or [])
    recent_and_revoked = list(recent) + [
        Contract(
            financing_type=c.financing_type,
            creditor=f"[Отозванный] {c.creditor}",
            contract_amount=c.contract_amount,
            periodic_payment=c.periodic_payment,
            outstanding=c.outstanding,
            current_overdue_days=c.current_overdue_days,
            current_overdue_amount=c.current_overdue_amount,
            end_date=c.end_date,
            max_overdue_days=c.max_overdue_days,
            max_overdue_amount=c.max_overdue_amount,
            cessionary=c.cessionary,
            status=c.status,
        )
        for c in revoked_list
    ]

    for row in range(30, 40):
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            safe_set(ws, f"{col}{row}", None)
    for idx, c in enumerate(recent_and_revoked[:10], start=1):
        row = 29 + idx
        safe_set(ws, f"B{row}", c.creditor)
        safe_set(ws, f"D{row}", c.cessionary if has_real_cession(c) else None)
        safe_set(ws, f"F{row}", c.end_date)
        safe_set(ws, f"G{row}", to_cell_str(c.max_overdue_days, blank_zero=blank_zero))
        safe_set(ws, f"H{row}", to_cell_str(c.max_overdue_amount, blank_zero=blank_zero))

    # Completed (> 5 years): rows 46..55 (10 rows) — headers occupy rows 44-45
    for row in range(46, 56):
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            safe_set(ws, f"{col}{row}", None)
    for idx, c in enumerate(old[:10], start=1):
        row = 45 + idx
        safe_set(ws, f"B{row}", c.creditor)
        safe_set(ws, f"D{row}", c.cessionary if has_real_cession(c) else None)
        safe_set(ws, f"F{row}", c.end_date)
        safe_set(ws, f"G{row}", to_cell_str(c.max_overdue_days, blank_zero=blank_zero))
        safe_set(ws, f"H{row}", to_cell_str(c.max_overdue_amount, blank_zero=blank_zero))

    # --- A4 print setup for all sheets ---
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
    for _sname in wb.sheetnames:
        _ws = wb[_sname]
        if _ws.sheet_properties is None:
            _ws.sheet_properties = WorksheetProperties()
        if _ws.sheet_properties.pageSetUpPr is None:
            _ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        _ws.sheet_properties.pageSetUpPr.fitToPage = True
        _ws.page_setup.paperSize = 9          # A4
        _ws.page_setup.orientation = "portrait"
        _ws.page_setup.fitToWidth = 1         # 1 страница по ширине
        _ws.page_setup.fitToHeight = 1        # 1 страница по высоте
        _ws.page_setup.scale = None
        _ws.page_margins = PageMargins(
            left=0.39, right=0.39,
            top=0.39, bottom=0.39,
            header=0.2, footer=0.2,
        )
    # --- end A4 print setup ---

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ----------------------------
# Fill "Анкета" from pension PDF (compact, 3 cols × 8 rows, per-org СРЗП)
# FIXED: Only write to top-left cells of merged ranges
# ----------------------------

def fill_anketa_from_pension(output_path: Path, pension: Dict[str, Any], sort_rows: bool = False) -> None:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    wb = load_workbook(str(output_path))
    if "Анкета" not in wb.sheetnames:
        LOGGER.warning("В output файле нет листа 'Анкета' — пропускаю заполнение.")
        wb.save(str(output_path))
        return

    ws = wb["Анкета"]

    fio = pension.get("fio")
    iin = pension.get("iin")

    # ФИО → B5 (merged B5:J5), ИИН → B6 (merged B6:D6)
    if fio:
        safe_set(ws, "B5", fio)
    if iin:
        safe_set(ws, "B6", iin)

    rows: List[PensionRow] = pension.get("rows") or []

    # Clear pension area rows 29-38
    for r in range(29, 44):
        for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            safe_set(ws, f"{c}{r}", None)

    # Group by sender (organisation)
    groups = group_pension_by_org(rows)
    org_names = list(groups.keys())[:3]  # max 3 organisations

    # New template layout:
    # Row 29: org headers  (A29, D29, G29)
    # Row 30: col headers  Дата/Сумма/Период × 3
    # Rows 31-38: data (8 rows)
    # Row 38 last: СРЗП label in A38, value in B38

    # Column mapping: org_idx -> (org_cell, date_col, sum_col, period_col)
    col_map = [
        ("A29", "A", "B", "C"),   # org 1
        ("D29", "D", "E", "F"),   # org 2
        ("G29", "G", "H", "I"),   # org 3
    ]

    total_srzp = 0
    org_srzp_list: List[str] = []

    for idx, bin_key in enumerate(org_names):
        org_cell, date_col, sum_col, period_col = col_map[idx]

        # Получаем читаемое название по BIN
        org_display = get_org_display_name(bin_key, rows)
        safe_set(ws, org_cell, org_display)

        org_rows = groups[bin_key]
        if sort_rows:
            from datetime import datetime as _dt
            def dkey(rr: PensionRow) -> "_dt":
                try:
                    return _dt.strptime(rr.date, "%d.%m.%Y")
                except Exception:
                    return _dt.max
            org_rows = sorted(org_rows, key=dkey)

        # Write up to 12 rows starting at row 31
        vals: List[float] = []
        for i, rr in enumerate(org_rows[:12]):
            r = 31 + i
            safe_set(ws, f"{date_col}{r}", rr.date)
            if isinstance(rr.amount, int) and rr.amount > 0:
                safe_set(ws, f"{sum_col}{r}", str(rr.amount))
                vals.append(float(rr.amount))
            else:
                safe_set(ws, f"{sum_col}{r}", None)
            safe_set(ws, f"{period_col}{r}", rr.period)

        # СРЗП по каждой организации отдельно
        srzp = calc_pension_avg_vals(vals)
        total_srzp += srzp
        org_srzp_list.append(f"{org_display}: {srzp:,}")
        LOGGER.info("Пенсионка БИН=%s (%s): строк=%s СРЗП=%s",
                    bin_key, org_display, len(vals), srzp)

    # Строка СРЗП: A43 = метка, B43 = значение
    # Если несколько орг — показываем каждую + итог
    safe_set(ws, "A43", "СРЗП")
    if len(org_names) > 1:
        detail = " | ".join(org_srzp_list)
        safe_set(ws, "B43", f"{total_srzp:,}  ({detail})")
    else:
        safe_set(ws, "B43", total_srzp)

    wb.save(str(output_path))


# ----------------------------
# CLI
# ----------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Извлечение данных из PDF в Excel v11.1 (2 листа, компактная пенсионка)."
    )
    parser.add_argument("--pdf", required=True, help="Путь к кредитному PDF.")
    parser.add_argument("--template", required=True, help="Путь к Excel-шаблону.")
    parser.add_argument("--output", required=True, help="Куда сохранить результат (.xlsx).")

    parser.add_argument("--pension-pdf", default=None, help="Путь к пенсионному PDF (опционально).")
    parser.add_argument("--pension-sort", action="store_true", help="Сортировать пенсионные строки по дате.")

    parser.add_argument("--only-overdue", action="store_true",
                        help="Заполнять только договоры с просрочкой.")

    parser.add_argument("--debug", action="store_true", help="Подробные логи.")
    parser.add_argument("--strict", action="store_true", help="Падать с ошибкой при провале валидации.")

    parser.add_argument("--ocr", action="store_true", help="Включить OCR-fallback.")
    parser.add_argument("--ocr-lang", default="rus", help="Язык OCR.")
    parser.add_argument("--poppler-path", default=None)
    parser.add_argument("--tesseract-cmd", default=None)
    parser.add_argument("--tessdata-dir", default=None)

    parser.add_argument("--plumber-layout", action="store_true",
                        help="Использовать pdfplumber.extract_text(layout=True).")

    parser.add_argument("--blank-zero", action="store_true", help="Писать нули как пустые ячейки (по умолчанию).")
    parser.add_argument("--no-blank-zero", action="store_true", help="Писать нули как '0'.")
    parser.add_argument("--only-loans-active", action="store_true", help="В активных заполнять только 'Займ'.")

    parser.add_argument("--validate-only", action="store_true", help="Только лог/валидация, без записи Excel.")

    args = parser.parse_args()
    setup_logging(args.debug)

    pdf_path = resolve_input_path(args.pdf)
    template_path = resolve_input_path(args.template)
    output_path = resolve_input_path(args.output)
    pension_path = resolve_input_path(args.pension_pdf) if args.pension_pdf else None

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF не найден: {pdf_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    blank_zero = True
    if args.no_blank_zero:
        blank_zero = False

    credit_data = extract_credit_report(
        pdf_path=pdf_path,
        enable_ocr=args.ocr,
        ocr_lang=args.ocr_lang,
        poppler_path=args.poppler_path,
        tesseract_cmd=args.tesseract_cmd,
        tessdata_dir=args.tessdata_dir,
        plumber_layout=args.plumber_layout,
    )
    validate_credit_data(credit_data, strict=args.strict)

    LOGGER.info("CREDIT: PKR=%s generated_at=%s iin=%s",
                credit_data.get("pkr"), credit_data.get("generated_at"), mask_iin(credit_data.get("iin")))
    LOGGER.info("COUNTS(header)=%s", {k: v for k, v in (credit_data.get("counts") or {}).items() if v is not None})
    LOGGER.info("PARSED: active=%s completed_recent=%s completed_old=%s revoked=%s",
                len(credit_data.get("active") or []),
                len(credit_data.get("completed_recent") or []),
                len(credit_data.get("completed_old") or []),
                len(credit_data.get("revoked") or []))

    pension_data = None
    if pension_path:
        if not pension_path.exists():
            raise FileNotFoundError(f"Pension PDF не найден: {pension_path}")
        pension_data = parse_pension_pdf_tables(pension_path)
        validate_pension_data(pension_data, strict=args.strict)
        validate_cross_iin(credit_data.get("iin"), pension_data.get("iin"), strict=args.strict)

        groups = group_pension_by_org(pension_data.get("rows") or [])
        LOGGER.info("PENSION: fio=%r iin=%s rows=%s orgs=%s",
                    pension_data.get("fio"), mask_iin(pension_data.get("iin")),
                    len(pension_data.get("rows") or []), list(groups.keys()))

    if args.only_overdue:
        credit_data["active"] = [c for c in (credit_data.get("active") or []) if contract_has_any_overdue(c)]
        credit_data["completed_recent"] = [
            c for c in (credit_data.get("completed_recent") or [])
            if contract_has_any_overdue(c) or contract_has_cession(c)
        ]
        credit_data["completed_old"] = [
            c for c in (credit_data.get("completed_old") or [])
            if contract_has_any_overdue(c) or contract_has_cession(c)
        ]
        LOGGER.info("only-overdue ON: active=%s completed_recent=%s completed_old=%s",
                    len(credit_data["active"]),
                    len(credit_data["completed_recent"]), len(credit_data["completed_old"]))

    if args.validate_only:
        LOGGER.info("validate-only: запись Excel пропущена.")
        return 0

    write_output_legacy_excel(
        template_path=template_path,
        output_path=output_path,
        pkr=credit_data.get("pkr"),
        active=credit_data.get("active") or [],
        recent=credit_data.get("completed_recent") or [],
        old=credit_data.get("completed_old") or [],
        blank_zero=blank_zero,
        only_loans_active=args.only_loans_active,
    )
    LOGGER.info("Кредитная история записана")

    if pension_data:
        fill_anketa_from_pension(output_path, pension_data, sort_rows=args.pension_sort)
        LOGGER.info("Анкета (пенсионка) дописана")

    LOGGER.info("Готово. Файл сохранён: %s", output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
